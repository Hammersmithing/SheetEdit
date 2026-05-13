#!/usr/bin/env python3
"""SheetEdit — lightweight .xlsx editor with call-sheet formatting primitives."""

import sys
import os
import json
import shutil
from pathlib import Path

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QTabWidget,
    QToolBar, QFileDialog, QColorDialog, QMessageBox, QVBoxLayout, QWidget,
    QLabel, QStatusBar, QMenu, QMenuBar, QSizePolicy, QStyledItemDelegate,
    QStyleOptionViewItem, QDialog, QHBoxLayout, QListWidget, QListWidgetItem,
    QPushButton, QLineEdit, QInputDialog, QGridLayout, QFrame, QPlainTextEdit,
    QComboBox, QScrollArea, QSplitter,
)
from PySide6.QtPrintSupport import QPrinter, QPrintDialog, QPrintPreviewDialog
from PySide6.QtGui import QPageLayout
from PySide6.QtCore import Qt, QSize, QRect, QRectF, QModelIndex, QPointF, QTimer
from PySide6.QtGui import (
    QAction, QColor, QFont, QBrush, QIcon, QKeySequence, QPainter, QPen,
    QShortcut, QFontMetrics,
)

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries
from openpyxl.styles import Font as XlFont, PatternFill, Alignment, Border, Side


# ── Theme + Color Handling ───────────────────────────────────────────────────

# Office default theme colors (Office 2007+ "Office" theme)
THEME_COLORS = [
    "#FFFFFF",  # 0  lt1 (background)
    "#000000",  # 1  dk1 (text)
    "#44546A",  # 2  dk2
    "#E7E6E6",  # 3  lt2
    "#4472C4",  # 4  accent1
    "#ED7D31",  # 5  accent2
    "#A5A5A5",  # 6  accent3
    "#FFC000",  # 7  accent4
    "#5B9BD5",  # 8  accent5
    "#70AD47",  # 9  accent6
    "#0563C1",  # 10 hyperlink
    "#954F72",  # 11 followed hyperlink
]


def _apply_tint(color: QColor, tint: float) -> QColor:
    """Apply Excel tint (-1.0 to 1.0) to a QColor. Positive = lighter, negative = darker."""
    r, g, b = color.red(), color.green(), color.blue()
    if tint > 0:
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)
    elif tint < 0:
        factor = 1.0 + tint  # e.g. tint=-0.5 → factor=0.5
        r = int(r * factor)
        g = int(g * factor)
        b = int(b * factor)
    return QColor(min(255, max(0, r)), min(255, max(0, g)), min(255, max(0, b)))


def xl_color_to_qcolor(xl_color, fallback=None):
    """Convert openpyxl color to QColor with full theme + tint support."""
    if xl_color is None:
        return fallback

    # Direct RGB
    if xl_color.type == "rgb" and xl_color.rgb and xl_color.rgb != "00000000":
        rgb = xl_color.rgb
        if len(rgb) == 8:
            qc = QColor(f"#{rgb[2:]}")
        else:
            qc = QColor(f"#{rgb}")
        if xl_color.tint and xl_color.tint != 0:
            qc = _apply_tint(qc, xl_color.tint)
        return qc

    # Theme color
    if xl_color.theme is not None:
        try:
            idx = xl_color.theme.__index__() if hasattr(xl_color.theme, '__index__') else int(str(xl_color.theme))
            if 0 <= idx < len(THEME_COLORS):
                qc = QColor(THEME_COLORS[idx])
            else:
                qc = QColor("#000000")
            if xl_color.tint and xl_color.tint != 0:
                tint_val = float(str(xl_color.tint)) if not isinstance(xl_color.tint, (int, float)) else xl_color.tint
                qc = _apply_tint(qc, tint_val)
            return qc
        except (ValueError, TypeError):
            return fallback

    # Indexed color (legacy)
    if xl_color.type == "indexed" and xl_color.indexed is not None:
        # Common indexed colors
        indexed_colors = {
            0: "#000000", 1: "#FFFFFF", 2: "#FF0000", 3: "#00FF00",
            4: "#0000FF", 5: "#FFFF00", 6: "#FF00FF", 7: "#00FFFF",
            8: "#000000", 9: "#FFFFFF", 10: "#FF0000", 11: "#00FF00",
            12: "#0000FF", 13: "#FFFF00", 14: "#FF00FF", 15: "#00FFFF",
            64: "#000000",  # system foreground
        }
        hex_val = indexed_colors.get(xl_color.indexed)
        if hex_val:
            return QColor(hex_val)

    return fallback


def qcolor_to_xl_rgb(qc: QColor) -> str:
    """QColor -> openpyxl hex string like 'FF4472C4'."""
    return f"FF{qc.red():02X}{qc.green():02X}{qc.blue():02X}"


HALIGN_MAP = {"left": Qt.AlignLeft, "center": Qt.AlignHCenter, "right": Qt.AlignRight}
VALIGN_MAP = {"top": Qt.AlignTop, "center": Qt.AlignVCenter, "bottom": Qt.AlignBottom}


# ── Templates ────────────────────────────────────────────────────────────────

TEMPLATES_DIR = Path.home() / ".sheetedit" / "templates"
SNIPPETS_DIR = Path.home() / ".sheetedit" / "snippets"
IMPORT_GUIDE_PATH = Path.home() / ".sheetedit" / "import_guide.md"


def _ensure_templates_dir():
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)


def _ensure_snippets_dir():
    SNIPPETS_DIR.mkdir(parents=True, exist_ok=True)


def _builtin_templates():
    """Return dict of {name: builder_function} for built-in templates."""
    templates = {}

    def _budget():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Budget"
        headers = ["Category", "Budgeted", "Actual", "Difference"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = XlFont(bold=True, size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = XlFont(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        categories = ["Housing", "Food", "Transport", "Utilities", "Entertainment", "Savings", "Other"]
        for r, cat in enumerate(categories, 2):
            ws.cell(row=r, column=1, value=cat)
            ws.cell(row=r, column=2, value=0)
            ws.cell(row=r, column=3, value=0)
            ws.cell(row=r, column=4).value = f"=B{r}-C{r}"
        total_row = len(categories) + 2
        ws.cell(row=total_row, column=1, value="Total").font = XlFont(bold=True)
        for c in range(2, 5):
            col_letter = get_column_letter(c)
            ws.cell(row=total_row, column=c).value = f"=SUM({col_letter}2:{col_letter}{total_row-1})"
            ws.cell(row=total_row, column=c).font = XlFont(bold=True)
        ws.column_dimensions["A"].width = 18
        for col in ["B", "C", "D"]:
            ws.column_dimensions[col].width = 14
        return wb

    def _invoice():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"
        ws.merge_cells("A1:D1")
        ws.cell(row=1, column=1, value="INVOICE").font = XlFont(bold=True, size=20)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        labels = [("Invoice #:", ""), ("Date:", ""), ("Bill To:", ""), ("", "")]
        for r, (label, val) in enumerate(labels, 3):
            ws.cell(row=r, column=1, value=label).font = XlFont(bold=True)
            ws.cell(row=r, column=2, value=val)
        header_row = 8
        headers = ["Description", "Qty", "Unit Price", "Amount"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=c, value=h)
            cell.font = XlFont(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for r in range(9, 14):
            ws.cell(row=r, column=4).value = f"=B{r}*C{r}"
        ws.cell(row=15, column=3, value="Total:").font = XlFont(bold=True)
        ws.cell(row=15, column=4).value = "=SUM(D9:D13)"
        ws.cell(row=15, column=4).font = XlFont(bold=True)
        ws.column_dimensions["A"].width = 30
        for col in ["B", "C", "D"]:
            ws.column_dimensions[col].width = 14
        return wb

    def _call_sheet():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Call Sheet"
        ws.merge_cells("A1:F1")
        ws.cell(row=1, column=1, value="CALL SHEET").font = XlFont(bold=True, size=18)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        info = ["Production:", "Date:", "Location:", "Call Time:"]
        for r, label in enumerate(info, 3):
            ws.cell(row=r, column=1, value=label).font = XlFont(bold=True)
        header_row = 8
        headers = ["Name", "Role", "Call Time", "Wrap Time", "Phone", "Notes"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=c, value=h)
            cell.font = XlFont(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for col in ["A", "B"]:
            ws.column_dimensions[col].width = 20
        for col in ["C", "D"]:
            ws.column_dimensions[col].width = 12
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 24
        return wb

    def _weekly_schedule():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Schedule"
        ws.cell(row=1, column=1, value="Time")
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        for c, day in enumerate(days, 2):
            cell = ws.cell(row=1, column=c, value=day)
            cell.font = XlFont(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        ws.cell(row=1, column=1).font = XlFont(bold=True, color="FFFFFF")
        ws.cell(row=1, column=1).fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        times = [f"{h}:00 {'AM' if h < 12 else 'PM'}" for h in range(7, 22)]
        for r, t in enumerate(times, 2):
            ws.cell(row=r, column=1, value=t).font = XlFont(size=10)
        ws.column_dimensions["A"].width = 12
        for c in range(2, 9):
            ws.column_dimensions[get_column_letter(c)].width = 16
        return wb

    templates["Budget"] = _budget
    templates["Invoice"] = _invoice
    templates["Call Sheet"] = _call_sheet
    templates["Weekly Schedule"] = _weekly_schedule
    return templates


BUILTIN_TEMPLATES = _builtin_templates()


# ── Column/Row Size Conversion ───────────────────────────────────────────────

def xl_col_width_to_px(width, default_font_size=11):
    """Convert Excel column width (character units) to pixels.
    Excel formula: pixels = Truncate(((width * max_digit_width + 5) / max_digit_width * 256) / 256 * max_digit_width)
    Simplified: for Calibri 11pt, max_digit_width ~ 7px. For Arial 11pt ~ 7px.
    """
    if width is None or width <= 0:
        return 64  # default
    # Excel standard: width is in character units of the default font
    # At 11pt, one character ~ 7.5 pixels, plus 5px padding on each side
    max_digit_w = 7.0 + (default_font_size - 11) * 0.5
    px = int(width * max_digit_w + 10)  # 5px padding each side
    return max(px, 20)


def xl_row_height_to_px(height):
    """Convert Excel row height (points) to pixels. 1pt = 1.333px at 96 DPI."""
    if height is None or height <= 0:
        return 21  # default row height
    return int(height * 96.0 / 72.0)


def px_to_xl_col_width(px, default_font_size=11):
    """Reverse of xl_col_width_to_px."""
    max_digit_w = 7.0 + (default_font_size - 11) * 0.5
    return max(0, (px - 10) / max_digit_w)


def px_to_xl_row_height(px):
    """Reverse of xl_row_height_to_px."""
    return px * 72.0 / 96.0


# ── Border Data + Delegate ───────────────────────────────────────────────────

# Border style → pen width and style
BORDER_STYLES = {
    "thin": (1, Qt.SolidLine),
    "medium": (2, Qt.SolidLine),
    "thick": (3, Qt.SolidLine),
    "dashed": (1, Qt.DashLine),
    "dotted": (1, Qt.DotLine),
    "double": (1, Qt.SolidLine),  # approximate
    "hair": (1, Qt.DotLine),
    "mediumDashed": (2, Qt.DashLine),
    "dashDot": (1, Qt.DashDotLine),
    "mediumDashDot": (2, Qt.DashDotLine),
    "dashDotDot": (1, Qt.DashDotDotLine),
    "mediumDashDotDot": (2, Qt.DashDotDotLine),
}


def _side_to_pen(side):
    """Convert openpyxl Side to (color_hex, width, style) or None."""
    if side is None or side.style is None or side.style == "none":
        return None
    color = "#000000"
    if side.color:
        qc = xl_color_to_qcolor(side.color)
        if qc and qc.isValid():
            color = qc.name()
    width, style = BORDER_STYLES.get(side.style, (1, Qt.SolidLine))
    return (color, width, style)


# Store border info as a role on QTableWidgetItem
BORDER_ROLE = Qt.UserRole + 100  # dict: {top, bottom, left, right} → (color, width, style)
RULE_ROLE = Qt.UserRole + 101    # bool: True if cell is in a ruled range


class BorderDelegate(QStyledItemDelegate):
    """Custom delegate that paints cell borders from stored border data."""

    def paint(self, painter: QPainter, option: QStyleOptionViewItem, index: QModelIndex):
        # Draw default content first
        super().paint(painter, option, index)

        # Only check custom roles if the cell has data
        try:
            border_data = index.data(BORDER_ROLE)
            rule_data = index.data(RULE_ROLE)
        except Exception:
            return

        if not border_data and not rule_data:
            return

        painter.save()
        rect = option.rect

        # Draw borders
        if border_data and isinstance(border_data, dict):
            for side_name, edge in [
                ("top", (rect.left(), rect.top(), rect.right(), rect.top())),
                ("bottom", (rect.left(), rect.bottom(), rect.right(), rect.bottom())),
                ("left", (rect.left(), rect.top(), rect.left(), rect.bottom())),
                ("right", (rect.right(), rect.top(), rect.right(), rect.bottom())),
            ]:
                info = border_data.get(side_name)
                if info:
                    color_hex, width, style = info
                    pen = QPen(QColor(color_hex), width, style)
                    painter.setPen(pen)
                    painter.drawLine(*edge)

        # Draw blue dot for ruled cells
        if rule_data:
            painter.setBrush(QBrush(QColor("#1A73E8")))
            painter.setPen(Qt.NoPen)
            painter.drawEllipse(rect.right() - 7, rect.top() + 2, 5, 5)

        painter.restore()


# ── Cell Conversion Helpers ──────────────────────────────────────────────────

def _xl_cell_to_item(cell):
    """Convert an openpyxl cell to a QTableWidgetItem."""
    item = QTableWidgetItem()

    # Value
    val = cell.value
    if val is not None:
        item.setText(str(val))

    # Font
    qf = QFont()
    if cell.font:
        qf.setFamily(cell.font.name or "Arial")
        qf.setPointSize(cell.font.size or 11)
        qf.setBold(cell.font.bold or False)
        qf.setItalic(cell.font.italic or False)
        qf.setUnderline(
            cell.font.underline is not None
            and cell.font.underline != "none"
        )
        fc = xl_color_to_qcolor(cell.font.color, QColor("#202124"))
        item.setForeground(QBrush(fc if fc else QColor("#202124")))
    else:
        item.setForeground(QBrush(QColor("#202124")))
    item.setFont(qf)

    # Fill
    has_fill = False
    if cell.fill and cell.fill.fgColor:
        bg = xl_color_to_qcolor(cell.fill.fgColor)
        if bg and bg.isValid() and bg != QColor(0, 0, 0):
            item.setBackground(QBrush(bg))
            has_fill = True
    if not has_fill:
        item.setBackground(QBrush(QColor("#FFFFFF")))

    # Alignment
    flags = Qt.AlignLeft | Qt.AlignVCenter
    if cell.alignment:
        h = HALIGN_MAP.get(cell.alignment.horizontal, Qt.AlignLeft)
        v = VALIGN_MAP.get(cell.alignment.vertical, Qt.AlignVCenter)
        flags = h | v
        if cell.alignment.wrap_text:
            flags |= Qt.TextWordWrap
    item.setTextAlignment(flags)

    # Borders
    if cell.border:
        bd = {}
        for side_name in ("top", "bottom", "left", "right"):
            side = getattr(cell.border, side_name, None)
            pen_info = _side_to_pen(side)
            if pen_info:
                bd[side_name] = pen_info
        if bd:
            item.setData(BORDER_ROLE, bd)

    return item


def _item_to_xl_cell(item, cell):
    """Write a QTableWidgetItem's data into an openpyxl cell."""
    if item is None:
        cell.value = None
        return

    cell.value = item.text() or None

    # Font
    qf = item.font()
    fg_brush = item.foreground()
    fg_color = (
        fg_brush.color()
        if fg_brush != QBrush()
        else QColor("#202124")
    )
    cell.font = XlFont(
        name=qf.family(),
        size=qf.pointSize(),
        bold=qf.bold(),
        italic=qf.italic(),
        underline="single" if qf.underline() else None,
        color=qcolor_to_xl_rgb(fg_color),
    )

    # Fill
    bg_brush = item.background()
    if bg_brush != QBrush() and bg_brush.color().isValid():
        bgc = bg_brush.color()
        if bgc != QColor("#FFFFFF"):
            cell.fill = PatternFill(
                start_color=qcolor_to_xl_rgb(bgc),
                end_color=qcolor_to_xl_rgb(bgc),
                fill_type="solid",
            )
        else:
            cell.fill = PatternFill(fill_type=None)
    else:
        cell.fill = PatternFill(fill_type=None)

    # Alignment
    flags = item.textAlignment()
    ha = "left"
    if flags & Qt.AlignHCenter:
        ha = "center"
    elif flags & Qt.AlignRight:
        ha = "right"
    va = "center"
    if flags & Qt.AlignTop:
        va = "top"
    elif flags & Qt.AlignBottom:
        va = "bottom"
    wrap = bool(flags & Qt.TextWordWrap)
    cell.alignment = Alignment(
        horizontal=ha, vertical=va, wrap_text=wrap
    )

    # Borders
    bd = item.data(BORDER_ROLE)
    if bd:
        sides = {}
        for side_name in ("top", "bottom", "left", "right"):
            info = bd.get(side_name)
            if info:
                color_hex, width, _ = info
                style = "thin"
                if width >= 3:
                    style = "thick"
                elif width >= 2:
                    style = "medium"
                sides[side_name] = Side(
                    style=style,
                    color=qcolor_to_xl_rgb(QColor(color_hex)),
                )
            else:
                sides[side_name] = Side(style=None)
        cell.border = Border(**sides)


# ── Snippet helpers ─────────────────────────────────────────────────────────

def save_snippet(name, sv, r1, c1, r2, c2):
    """Save the selected rectangle from a SheetView as a snippet .xlsx."""
    _ensure_snippets_dir()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Snippet"

    rows = r2 - r1 + 1
    cols = c2 - c1 + 1

    for ri in range(rows):
        for ci in range(cols):
            item = sv.item(r1 + ri, c1 + ci)
            dest_cell = ws.cell(row=ri + 1, column=ci + 1)
            _item_to_xl_cell(item, dest_cell)

    # Column widths
    for ci in range(cols):
        col_letter = get_column_letter(ci + 1)
        ws.column_dimensions[col_letter].width = px_to_xl_col_width(
            sv.columnWidth(c1 + ci)
        )

    # Row heights
    for ri in range(rows):
        ws.row_dimensions[ri + 1].height = px_to_xl_row_height(
            sv.rowHeight(r1 + ri)
        )

    # Merges offset to (0,0)
    for m in sv.merges:
        mr1, mc1, mr2, mc2 = m
        # Check overlap with selection
        if mr1 >= r1 and mr2 <= r2 and mc1 >= c1 and mc2 <= c2:
            ws.merge_cells(
                start_row=mr1 - r1 + 1, start_column=mc1 - c1 + 1,
                end_row=mr2 - r1 + 1, end_column=mc2 - c1 + 1,
            )

    wb.save(str(SNIPPETS_DIR / f"{name}.xlsx"))


def insert_snippet(name, sv, dest_r, dest_c):
    """Load a snippet .xlsx and insert its cells at (dest_r, dest_c) in sv."""
    path = SNIPPETS_DIR / f"{name}.xlsx"
    if not path.exists():
        return
    # currentRow/currentColumn return -1 when no cell is selected (e.g. fresh
    # workbook). Clamp to 0 so the snippet anchors at A1 instead of silently
    # dropping the first row/column.
    if dest_r < 0:
        dest_r = 0
    if dest_c < 0:
        dest_c = 0
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active

    rows = ws.max_row or 1
    cols = ws.max_column or 1

    # Expand sheet if needed
    needed_rows = dest_r + rows
    needed_cols = dest_c + cols
    if sv.rowCount() < needed_rows:
        sv.setRowCount(needed_rows)
    if sv.columnCount() < needed_cols:
        sv.setColumnCount(needed_cols)
        sv.setHorizontalHeaderLabels(
            [get_column_letter(c + 1) for c in range(sv.columnCount())]
        )

    # Collect coords/cols/rows the insertion will touch, for a single
    # "snippet_insert" undo entry that reverses the whole operation atomically.
    cell_coords = [(dest_r + ri, dest_c + ci)
                   for ri in range(rows) for ci in range(cols)]
    touched_cols = [dest_c + ci for ci in range(cols)
                    if get_column_letter(ci + 1) in ws.column_dimensions
                    and ws.column_dimensions[get_column_letter(ci + 1)].width]
    touched_rows = [dest_r + ri for ri in range(rows)
                    if (ri + 1) in ws.row_dimensions
                    and ws.row_dimensions[ri + 1].height]

    before_cells = sv._snapshot_cells(cell_coords)
    before_col_widths = {c: sv.columnWidth(c) for c in touched_cols}
    before_row_heights = {r: sv.rowHeight(r) for r in touched_rows}

    # Batch updates to avoid rendering glitches
    sv.setUpdatesEnabled(False)
    sv.blockSignals(True)

    # Insert cells
    for ri in range(rows):
        for ci in range(cols):
            cell = ws.cell(row=ri + 1, column=ci + 1)
            item = _xl_cell_to_item(cell)
            sv.setItem(dest_r + ri, dest_c + ci, item)

    # Column widths
    for ci in range(cols):
        col_letter = get_column_letter(ci + 1)
        if col_letter in ws.column_dimensions:
            w = ws.column_dimensions[col_letter].width
            if w:
                sv.setColumnWidth(dest_c + ci, xl_col_width_to_px(w))

    # Row heights
    for ri in range(rows):
        if (ri + 1) in ws.row_dimensions:
            h = ws.row_dimensions[ri + 1].height
            if h:
                sv.setRowHeight(dest_r + ri, xl_row_height_to_px(h))

    # Merges offset to destination — track them for undo
    merges_added = []
    for rng in ws.merged_cells.ranges:
        mr1, mc1 = rng.min_row - 1, rng.min_col - 1
        mr2, mc2 = rng.max_row - 1, rng.max_col - 1
        new_r1 = dest_r + mr1
        new_c1 = dest_c + mc1
        new_r2 = dest_r + mr2
        new_c2 = dest_c + mc2
        sv.setSpan(new_r1, new_c1, new_r2 - new_r1 + 1, new_c2 - new_c1 + 1)
        sv.merges.append((new_r1, new_c1, new_r2, new_c2))
        merges_added.append((new_r1, new_c1, new_r2, new_c2))

    sv.blockSignals(False)
    sv.setUpdatesEnabled(True)
    sv.viewport().update()

    after_cells = sv._snapshot_cells(cell_coords)
    after_col_widths = {c: sv.columnWidth(c) for c in touched_cols}
    after_row_heights = {r: sv.rowHeight(r) for r in touched_rows}

    sv._push_undo_entry(("snippet_insert", {
        "before_cells": before_cells,
        "after_cells": after_cells,
        "merges_added": merges_added,
        "before_col_widths": before_col_widths,
        "after_col_widths": after_col_widths,
        "before_row_heights": before_row_heights,
        "after_row_heights": after_row_heights,
    }))
    win = sv.window()
    if hasattr(win, '_dirty'):
        win._dirty = True


def list_snippets():
    """Return sorted list of snippet names."""
    _ensure_snippets_dir()
    return sorted(p.stem for p in SNIPPETS_DIR.glob("*.xlsx"))


def delete_snippet(name):
    """Delete a snippet file."""
    path = SNIPPETS_DIR / f"{name}.xlsx"
    if path.exists():
        path.unlink()


def _load_snippet_meta(name):
    """Load a snippet's meta.json. Returns dict or None."""
    path = SNIPPETS_DIR / f"{name}.meta.json"
    if path.exists():
        return json.loads(path.read_text())
    return None


def _copy_cell(src, dst):
    """Copy value and formatting from one openpyxl cell to another."""
    dst.value = src.value
    if src.font:
        dst.font = src.font.copy()
    if src.fill:
        dst.fill = src.fill.copy()
    if src.alignment:
        dst.alignment = src.alignment.copy()
    if src.border:
        dst.border = src.border.copy()
    if src.number_format:
        dst.number_format = src.number_format


def _copy_row_formatting(snip_ws, template_row, dst_ws, dst_row, num_cols):
    """Copy formatting (not values) from a template row to a destination row."""
    for ci in range(1, num_cols + 1):
        src = snip_ws.cell(row=template_row, column=ci)
        dst = dst_ws.cell(row=dst_row, column=ci)
        if src.font:
            dst.font = src.font.copy()
        if src.fill:
            dst.fill = src.fill.copy()
        if src.alignment:
            dst.alignment = src.alignment.copy()
        if src.border:
            dst.border = src.border.copy()
        if src.number_format:
            dst.number_format = src.number_format
    # Copy row height
    if template_row in snip_ws.row_dimensions:
        h = snip_ws.row_dimensions[template_row].height
        if h:
            dst_ws.row_dimensions[dst_row].height = h


def _build_snippet_to_ws(ws, snip_name, current_row, data_count=None):
    """Write a snippet into ws starting at current_row. Returns rows written.

    If data_count is provided and the snippet has a template_row in its meta,
    the repeating section expands to data_count rows (cloning the template
    row's formatting). Footer rows shift down accordingly.
    """
    path = SNIPPETS_DIR / f"{snip_name}.xlsx"
    if not path.exists():
        return 0

    snip_wb = openpyxl.load_workbook(str(path))
    snip_ws = snip_wb.active
    snip_rows = snip_ws.max_row or 1
    snip_cols = snip_ws.max_column or 1
    meta = _load_snippet_meta(snip_name)

    if meta and meta.get("template_row") and data_count is not None:
        template_row = meta["template_row"]
        header_rows = meta.get("header_rows", [])
        footer_rows = meta.get("footer_rows", [])
        # Determine how many placeholder rows exist between header and footer
        placeholder_start = template_row
        placeholder_end = max(
            r for r in range(1, snip_rows + 1)
            if r not in header_rows and r not in footer_rows
        )
        num_placeholders = placeholder_end - placeholder_start + 1
        actual_data_rows = max(data_count, 1)

        dest_row = current_row
        # 1. Copy header rows
        for hr in header_rows:
            for ci in range(1, snip_cols + 1):
                _copy_cell(snip_ws.cell(row=hr, column=ci),
                           ws.cell(row=dest_row + hr, column=ci))
            if hr in snip_ws.row_dimensions:
                h = snip_ws.row_dimensions[hr].height
                if h:
                    ws.row_dimensions[dest_row + hr].height = h
        dest_row_data_start = current_row + len(header_rows)

        # 2. Create data rows using template row formatting
        for di in range(actual_data_rows):
            row_num = dest_row_data_start + di + 1
            _copy_row_formatting(snip_ws, template_row, ws, row_num, snip_cols)

        dest_row_after_data = dest_row_data_start + actual_data_rows

        # 3. Copy footer rows
        for fi, fr in enumerate(footer_rows):
            footer_dest = dest_row_after_data + fi + 1
            for ci in range(1, snip_cols + 1):
                _copy_cell(snip_ws.cell(row=fr, column=ci),
                           ws.cell(row=footer_dest, column=ci))
            if fr in snip_ws.row_dimensions:
                h = snip_ws.row_dimensions[fr].height
                if h:
                    ws.row_dimensions[footer_dest].height = h

        total_rows = len(header_rows) + actual_data_rows + len(footer_rows)

        # Copy column widths
        for ci in range(snip_cols):
            col_letter = get_column_letter(ci + 1)
            if col_letter in snip_ws.column_dimensions:
                src_w = snip_ws.column_dimensions[col_letter].width
                if src_w:
                    existing = ws.column_dimensions[col_letter].width
                    if not existing or src_w > existing:
                        ws.column_dimensions[col_letter].width = src_w

        # Copy merges from header rows (offset to current_row)
        for rng in snip_ws.merged_cells.ranges:
            if rng.min_row in header_rows and rng.max_row in header_rows:
                ws.merge_cells(
                    start_row=rng.min_row + current_row,
                    start_column=rng.min_col,
                    end_row=rng.max_row + current_row,
                    end_column=rng.max_col,
                )
        # Copy merges from footer rows (offset to new position)
        footer_offset = dest_row_after_data - footer_rows[0] + 1 if footer_rows else 0
        for rng in snip_ws.merged_cells.ranges:
            if rng.min_row in footer_rows:
                ws.merge_cells(
                    start_row=rng.min_row + footer_offset,
                    start_column=rng.min_col,
                    end_row=rng.max_row + footer_offset,
                    end_column=rng.max_col,
                )

        return total_rows
    else:
        # Static copy — no expansion (original behavior)
        for ri in range(snip_rows):
            for ci in range(snip_cols):
                _copy_cell(snip_ws.cell(row=ri + 1, column=ci + 1),
                           ws.cell(row=current_row + ri + 1, column=ci + 1))

        for ci in range(snip_cols):
            col_letter = get_column_letter(ci + 1)
            if col_letter in snip_ws.column_dimensions:
                src_w = snip_ws.column_dimensions[col_letter].width
                if src_w:
                    existing = ws.column_dimensions[col_letter].width
                    if not existing or src_w > existing:
                        ws.column_dimensions[col_letter].width = src_w

        for ri in range(snip_rows):
            if (ri + 1) in snip_ws.row_dimensions:
                h = snip_ws.row_dimensions[ri + 1].height
                if h:
                    ws.row_dimensions[current_row + ri + 1].height = h

        for rng in snip_ws.merged_cells.ranges:
            ws.merge_cells(
                start_row=rng.min_row + current_row,
                start_column=rng.min_col,
                end_row=rng.max_row + current_row,
                end_column=rng.max_col,
            )

        return snip_rows


# ── SheetView — one tab ─────────────────────────────────────────────────────

class SheetView(QTableWidget):
    """Displays one openpyxl worksheet in a QTableWidget."""

    def __init__(self, ws, parent=None):
        super().__init__(parent)
        self.ws = ws
        self.merges = []
        self._undo_stack = []
        self._redo_stack = []
        self._MAX_UNDO = 500
        self.setItemDelegate(BorderDelegate(self))
        self.setShowGrid(True)  # light grid from stylesheet, borders drawn on top
        # Google Sheets-style: typing goes straight into cells,
        # double-click or F2 for inline editor
        self.setEditTriggers(QTableWidget.DoubleClicked)
        self._editing = False  # True when user is typing into a cell
        # Clipboard state for copy/cut with marching ants
        self._clip_cells = []  # list of (row, col, snapshot_data)
        self._clip_mode = None  # "copy" or "cut"
        self._clip_range = None  # (r1, c1, r2, c2) for drawing border
        self._march_timer = None  # timer for marching ants animation
        self._march_offset = 0
        self._load()
        self.itemChanged.connect(self._on_item_changed)
        self.currentCellChanged.connect(self._cell_moved)
        self._tracking_edits = True
        # Track column/row resize for undo (debounced so one drag = one undo)
        self.horizontalHeader().sectionResized.connect(self._col_resized)
        self.verticalHeader().sectionResized.connect(self._row_resized)
        self._resize_tracking = True
        self._resize_pending = None  # ("col"|"row", index, original_size)
        self._resize_timer = QTimer()
        self._resize_timer.setSingleShot(True)
        self._resize_timer.setInterval(300)
        self._resize_timer.timeout.connect(self._flush_resize)

    def _snapshot_cells(self, cells):
        """Capture current state of a list of (row, col) cells."""
        snap = {}
        for r, c in cells:
            item = self.item(r, c)
            if item:
                snap[(r, c)] = {
                    "text": item.text(),
                    "fg": item.foreground().color() if item.foreground().style() != Qt.NoBrush else None,
                    "bg": item.background().color() if item.background().style() != Qt.NoBrush else None,
                    "font": QFont(item.font()),
                    "align": item.textAlignment(),
                    "border": item.data(BORDER_ROLE),
                }
            else:
                snap[(r, c)] = None
        return snap

    def _restore_snapshot(self, snap):
        """Restore cells from a snapshot dict."""
        self._tracking_edits = False
        for (r, c), data in snap.items():
            if data is None:
                # Cell didn't exist before — clear it fully
                item = self.item(r, c)
                if item:
                    item.setText("")
                    item.setForeground(QBrush(QColor("#202124")))
                    item.setBackground(QBrush(QColor("#FFFFFF")))
                    item.setFont(QFont("Arial", 11))
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    item.setData(BORDER_ROLE, None)
                continue
            item = self.item(r, c)
            if not item:
                item = QTableWidgetItem()
                self.setItem(r, c, item)
            item.setText(data["text"])
            item.setForeground(QBrush(data["fg"]) if data["fg"] else QBrush(QColor("#202124")))
            item.setBackground(QBrush(data["bg"]) if data["bg"] else QBrush(QColor("#FFFFFF")))
            item.setFont(data["font"])
            item.setTextAlignment(data["align"])
            item.setData(BORDER_ROLE, data["border"])
        self._tracking_edits = True

    def push_undo(self, cells):
        """Save current state of cells for undo. Call BEFORE making changes."""
        snap = self._snapshot_cells(cells)
        self._undo_stack.append(("cells", snap))
        if len(self._undo_stack) > self._MAX_UNDO:
            self._undo_stack.pop(0)
        self._redo_stack.clear()
        win = self.window()
        if hasattr(win, '_dirty'):
            win._dirty = True

    def _push_undo_entry(self, entry):
        """Push a raw undo entry (any type)."""
        self._undo_stack.append(entry)
        if len(self._undo_stack) > self._MAX_UNDO:
            self._undo_stack.pop(0)
        self._redo_stack.clear()

    def undo(self):
        self._flush_resize()
        if not self._undo_stack:
            return
        entry = self._undo_stack.pop()
        kind = entry[0]
        if kind == "cells":
            old_snap = entry[1]
            current = self._snapshot_cells(old_snap.keys())
            self._redo_stack.append(("cells", current))
            self._restore_snapshot(old_snap)
        elif kind == "col_resize":
            _, col, old_w = entry
            cur_w = self.columnWidth(col)
            self._redo_stack.append(("col_resize", col, cur_w))
            self._resize_tracking = False
            self.setColumnWidth(col, old_w)
            self._resize_tracking = True
        elif kind == "row_resize":
            _, row, old_h = entry
            cur_h = self.rowHeight(row)
            self._redo_stack.append(("row_resize", row, cur_h))
            self._resize_tracking = False
            self.setRowHeight(row, old_h)
            self._resize_tracking = True
        elif kind == "snippet_insert":
            self._apply_snippet_state(entry[1], direction="undo")
            self._redo_stack.append(entry)

    def redo(self):
        self._flush_resize()
        if not self._redo_stack:
            return
        entry = self._redo_stack.pop()
        kind = entry[0]
        if kind == "cells":
            new_snap = entry[1]
            current = self._snapshot_cells(new_snap.keys())
            self._undo_stack.append(("cells", current))
            self._restore_snapshot(new_snap)
        elif kind == "col_resize":
            _, col, new_w = entry
            cur_w = self.columnWidth(col)
            self._undo_stack.append(("col_resize", col, cur_w))
            self._resize_tracking = False
            self.setColumnWidth(col, new_w)
            self._resize_tracking = True
        elif kind == "row_resize":
            _, row, new_h = entry
            cur_h = self.rowHeight(row)
            self._undo_stack.append(("row_resize", row, cur_h))
            self._resize_tracking = False
            self.setRowHeight(row, new_h)
            self._resize_tracking = True
        elif kind == "snippet_insert":
            self._apply_snippet_state(entry[1], direction="redo")
            self._undo_stack.append(entry)

    def _apply_snippet_state(self, data, direction):
        """Reverse or replay a snippet insertion in one atomic step.

        direction="undo": restore pre-insert cells + remove the merges that
        the insertion added. direction="redo": restore post-insert cells +
        re-add those merges.
        """
        if direction == "undo":
            cells = data["before_cells"]
            col_widths = data["before_col_widths"]
            row_heights = data["before_row_heights"]
            remove_merges = True
        else:
            cells = data["after_cells"]
            col_widths = data["after_col_widths"]
            row_heights = data["after_row_heights"]
            remove_merges = False
        self.setUpdatesEnabled(False)
        self.blockSignals(True)
        self._restore_snapshot(cells)
        added = data["merges_added"]
        if remove_merges:
            for m in added:
                if m in self.merges:
                    self.merges.remove(m)
        else:
            for m in added:
                if m not in self.merges:
                    self.merges.append(m)
        self.clearSpans()
        for m in self.merges:
            if m[2] > m[0] or m[3] > m[1]:
                self.setSpan(m[0], m[1], m[2] - m[0] + 1, m[3] - m[1] + 1)
        self._resize_tracking = False
        for c, w in col_widths.items():
            self.setColumnWidth(c, w)
        for r, h in row_heights.items():
            self.setRowHeight(r, h)
        self._resize_tracking = True
        self.blockSignals(False)
        self.setUpdatesEnabled(True)
        self.viewport().update()

    def _col_resized(self, col, old_size, new_size):
        if not self._resize_tracking:
            return
        # Only capture the original size at the start of a drag
        if self._resize_pending is None or self._resize_pending[:2] != ("col", col):
            self._flush_resize()  # flush any previous pending resize
            self._resize_pending = ("col", col, old_size)
        self._resize_timer.start()

    def _row_resized(self, row, old_size, new_size):
        if not self._resize_tracking:
            return
        if self._resize_pending is None or self._resize_pending[:2] != ("row", row):
            self._flush_resize()
            self._resize_pending = ("row", row, old_size)
        self._resize_timer.start()

    def _flush_resize(self):
        """Commit the pending resize as a single undo entry."""
        if self._resize_pending is None:
            return
        kind, idx, orig = self._resize_pending
        self._resize_pending = None
        if kind == "col":
            self._push_undo_entry(("col_resize", idx, orig))
        else:
            self._push_undo_entry(("row_resize", idx, orig))

    def copy_cells(self):
        """Copy selected cells to internal clipboard."""
        sel = self.selectedRanges()
        if not sel:
            return
        r = sel[0]
        self._clip_cells = []
        for ri in range(r.topRow(), r.bottomRow() + 1):
            for ci in range(r.leftColumn(), r.rightColumn() + 1):
                snap = self._snapshot_cells([(ri, ci)])
                self._clip_cells.append((ri - r.topRow(), ci - r.leftColumn(), snap.get((ri, ci))))
        self._clip_mode = "copy"
        self._clip_range = (r.topRow(), r.leftColumn(), r.bottomRow(), r.rightColumn())
        self._start_marching_ants()

    def cut_cells(self):
        """Cut selected cells to internal clipboard."""
        self.copy_cells()
        self._clip_mode = "cut"

    def paste_cells(self):
        """Paste from internal clipboard to current position."""
        if not self._clip_cells:
            return
        dest_r, dest_c = self.currentRow(), self.currentColumn()
        if dest_r < 0 or dest_c < 0:
            return

        # Collect coords for undo — include both destination AND source (if cut)
        coords = [(dest_r + dr, dest_c + dc) for dr, dc, _ in self._clip_cells]
        if self._clip_mode == "cut" and self._clip_range:
            r1, c1, r2, c2 = self._clip_range
            for ri in range(r1, r2 + 1):
                for ci in range(c1, c2 + 1):
                    if (ri, ci) not in coords:
                        coords.append((ri, ci))
        self.push_undo(coords)

        self._tracking_edits = False
        for dr, dc, data in self._clip_cells:
            r, c = dest_r + dr, dest_c + dc
            if r >= self.rowCount() or c >= self.columnCount():
                continue
            if data is None:
                item = self.item(r, c)
                if item:
                    item.setText("")
            else:
                item = self._ensure_item(r, c)
                item.setText(data.get("text", ""))
                if data.get("fg"):
                    item.setForeground(QBrush(data["fg"]))
                if data.get("bg"):
                    item.setBackground(QBrush(data["bg"]))
                if data.get("font"):
                    item.setFont(data["font"])
                if data.get("align") is not None:
                    item.setTextAlignment(data["align"])
                if data.get("border"):
                    item.setData(BORDER_ROLE, data["border"])
        self._tracking_edits = True

        # If cut, clear the source cells
        if self._clip_mode == "cut":
            r1, c1, r2, c2 = self._clip_range
            self._tracking_edits = False
            for ri in range(r1, r2 + 1):
                for ci in range(c1, c2 + 1):
                    item = self.item(ri, ci)
                    if item:
                        item.setText("")
            self._tracking_edits = True
            self.cancel_clip()
        # Copy mode: keep clipboard active for multiple pastes

        win = self.window()
        if hasattr(win, '_dirty'):
            win._dirty = True

    def cancel_clip(self):
        """Cancel the current copy/cut operation."""
        self._clip_cells = []
        self._clip_mode = None
        self._clip_range = None
        if self._march_timer:
            self._march_timer.stop()
            self._march_timer = None
        self._march_offset = 0
        self.viewport().update()

    def _start_marching_ants(self):
        """Start the marching ants animation around clipped cells."""
        from PySide6.QtCore import QTimer
        if self._march_timer:
            self._march_timer.stop()
        self._march_offset = 0
        self._march_timer = QTimer(self)
        self._march_timer.timeout.connect(self._march_tick)
        self._march_timer.start(150)

    def _march_tick(self):
        self._march_offset = (self._march_offset + 1) % 8
        self.viewport().update()

    def paintEvent(self, event):
        """Draw marching ants border around copy/cut selection."""
        super().paintEvent(event)
        if not self._clip_range:
            return
        try:
            r1, c1, r2, c2 = self._clip_range
            rect_tl = self.visualRect(self.model().index(r1, c1))
            rect_br = self.visualRect(self.model().index(r2, c2))
            if rect_tl.isNull() or rect_br.isNull():
                return
            rect = rect_tl.united(rect_br)
            painter = QPainter(self.viewport())
            pen = QPen(QColor("#1A73E8"), 2, Qt.DashLine)
            pen.setDashOffset(self._march_offset)
            painter.setPen(pen)
            painter.drawRect(rect.adjusted(0, 0, -1, -1))
            painter.end()
        except Exception:
            pass

    def _cell_moved(self, row, col, prev_row, prev_col):
        self._editing = False

    def _on_item_changed(self, item):
        """Track direct cell edits by the user."""
        if not self._tracking_edits:
            return
        # For single cell edits, we push a minimal undo entry
        # (This captures typing; bulk operations push their own undo before changing)
        # Mark dirty, update formula bar and check rules
        win = self.window()
        if hasattr(win, '_dirty'):
            win._dirty = True
        if hasattr(win, '_update_formula_bar'):
            win._update_formula_bar()
        if hasattr(win, '_check_and_warn'):
            win._check_and_warn([(item.row(), item.column())])

    def _ensure_item(self, row, col):
        """Get or create a QTableWidgetItem at (row, col)."""
        item = self.item(row, col)
        if item is None:
            item = QTableWidgetItem()
            item.setForeground(QBrush(QColor("#202124")))
            item.setBackground(QBrush(QColor("#FFFFFF")))
            item.setFont(QFont("Arial", 11))
            item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.setItem(row, col, item)
        return item

    def keyPressEvent(self, event):
        key = event.key()
        mods = event.modifiers()

        # If Qt's inline editor is open (F2 / double-click), let it handle everything
        if self.state() == QTableWidget.EditingState:
            if key in (Qt.Key_Return, Qt.Key_Enter):
                super().keyPressEvent(event)
                row, col = self.currentRow(), self.currentColumn()
                if row < self.rowCount() - 1:
                    self.setCurrentCell(row + 1, col)
                self._editing = False
                return
            if key == Qt.Key_Escape:
                super().keyPressEvent(event)
                self._editing = False
                return
            super().keyPressEvent(event)
            return

        # Delete/Backspace: clear selected cells
        if key in (Qt.Key_Delete, Qt.Key_Backspace):
            if self._editing:
                # Backspace while typing: remove last char
                row, col = self.currentRow(), self.currentColumn()
                item = self.item(row, col)
                if item and item.text():
                    self._tracking_edits = False
                    item.setText(item.text()[:-1])
                    self._tracking_edits = True
                    win = self.window()
                    if hasattr(win, '_update_formula_bar'):
                        win._update_formula_bar()
                else:
                    self._editing = False
                return
            items = self.selectedItems()
            if items:
                coords = [(it.row(), it.column()) for it in items]
                self.push_undo(coords)
                self._tracking_edits = False
                for it in items:
                    it.setText("")
                self._tracking_edits = True
                win = self.window()
                if hasattr(win, '_update_formula_bar'):
                    win._update_formula_bar()
            return

        # Enter/Return: move down
        if key in (Qt.Key_Return, Qt.Key_Enter):
            self._editing = False
            row, col = self.currentRow(), self.currentColumn()
            if row < self.rowCount() - 1:
                self.setCurrentCell(row + 1, col)
            return

        # Tab / Shift+Tab: move right/left
        if key == Qt.Key_Tab:
            self._editing = False
            row, col = self.currentRow(), self.currentColumn()
            if mods & Qt.ShiftModifier:
                if col > 0:
                    self.setCurrentCell(row, col - 1)
            else:
                if col < self.columnCount() - 1:
                    self.setCurrentCell(row, col + 1)
            return

        # Escape: cancel clip or stop typing mode
        if key == Qt.Key_Escape:
            if self._clip_range:
                self.cancel_clip()
            self._editing = False
            return

        # Cmd/Ctrl+C: copy
        if key == Qt.Key_C and (mods & Qt.ControlModifier or mods & Qt.MetaModifier):
            self.copy_cells()
            return

        # Cmd/Ctrl+X: cut
        if key == Qt.Key_X and (mods & Qt.ControlModifier or mods & Qt.MetaModifier):
            self.cut_cells()
            return

        # Cmd/Ctrl+V: paste
        if key == Qt.Key_V and (mods & Qt.ControlModifier or mods & Qt.MetaModifier):
            self.paste_cells()
            return

        # Cmd/Ctrl+Z: undo
        if key == Qt.Key_Z and (mods & Qt.ControlModifier or mods & Qt.MetaModifier):
            if mods & Qt.ShiftModifier:
                self.redo()
            else:
                self.undo()
            return

        # F2: open Qt inline editor
        if key == Qt.Key_F2:
            self.editItem(self.currentItem())
            return

        # Arrow keys: navigate (stop typing mode)
        if key in (Qt.Key_Left, Qt.Key_Right, Qt.Key_Up, Qt.Key_Down):
            self._editing = False
            super().keyPressEvent(event)
            return

        # Printable text: type directly into cell
        text = event.text()
        if text and text.isprintable() and not (mods & Qt.ControlModifier) and not (mods & Qt.MetaModifier):
            row, col = self.currentRow(), self.currentColumn()
            if row < 0 or col < 0:
                return
            item = self._ensure_item(row, col)
            if not self._editing:
                # First keystroke: push undo and replace content
                self.push_undo([(row, col)])
                self._tracking_edits = False
                item.setText(text)
                self._tracking_edits = True
                self._editing = True
            else:
                # Subsequent keystrokes: append
                self._tracking_edits = False
                item.setText(item.text() + text)
                self._tracking_edits = True
            win = self.window()
            if hasattr(win, '_update_formula_bar'):
                win._update_formula_bar()
            return

        super().keyPressEvent(event)

    def contextMenuEvent(self, event):
        menu = QMenu(self)

        # Save Selection as Snippet
        save_act = QAction("Save Selection as Snippet...", self)
        save_act.triggered.connect(self._ctx_save_snippet)
        menu.addAction(save_act)

        # Insert Snippet submenu
        snippets = list_snippets()
        if snippets:
            sub = menu.addMenu("Insert Snippet")
            for name in snippets:
                act = QAction(name, self)
                act.triggered.connect(lambda checked, n=name: self._ctx_insert_snippet(n))
                sub.addAction(act)

        # Edit Import Rules submenu
        if snippets:
            rules_sub = menu.addMenu("Edit Import Rules")
            for name in snippets:
                act = QAction(name, self)
                act.triggered.connect(lambda checked, n=name: self._ctx_open_snippet_rules(n))
                rules_sub.addAction(act)

        # Manage Snippets
        manage_act = QAction("Manage Snippets...", self)
        manage_act.triggered.connect(self._ctx_manage_snippets)
        menu.addAction(manage_act)

        menu.exec(event.globalPos())

    def _ctx_save_snippet(self):
        win = self.window()
        rng = win._selected_range()
        if rng is None:
            QMessageBox.information(self, "Snippet", "Select a range of cells first.")
            return
        r1, c1, r2, c2 = rng
        name, ok = QInputDialog.getText(self, "Save Snippet", "Snippet name:")
        if not ok or not name.strip():
            return
        name = name.strip()
        dest = SNIPPETS_DIR / f"{name}.xlsx"
        if dest.exists():
            reply = QMessageBox.question(
                self, "Overwrite Snippet",
                f'Snippet "{name}" already exists. Overwrite?',
                QMessageBox.Yes | QMessageBox.No,
            )
            if reply != QMessageBox.Yes:
                return
        save_snippet(name, self, r1, c1, r2, c2)
        win.statusBar().showMessage(f"Saved snippet: {name}")

    def _ctx_insert_snippet(self, name):
        row = self.currentRow()
        col = self.currentColumn()
        insert_snippet(name, self, row, col)
        self.window().statusBar().showMessage(f"Inserted snippet: {name}")

    def _ctx_manage_snippets(self):
        snippets = list_snippets()
        if not snippets:
            QMessageBox.information(self, "Snippets", "No snippets saved yet.")
            return
        dlg = QDialog(self)
        dlg.setWindowTitle("Manage Snippets")
        dlg.setMinimumSize(300, 250)
        layout = QVBoxLayout(dlg)
        lw = QListWidget()
        for s in snippets:
            lw.addItem(s)
        layout.addWidget(lw)
        del_btn = QPushButton("Delete Selected")
        def _delete():
            item = lw.currentItem()
            if item:
                delete_snippet(item.text())
                lw.takeItem(lw.row(item))
        del_btn.clicked.connect(_delete)
        layout.addWidget(del_btn)
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dlg.accept)
        layout.addWidget(close_btn)
        dlg.exec()

    def _ctx_open_snippet_rules(self, snippet_name):
        """Open the visual snippet rules editor."""
        dlg = SnippetRulesEditor(snippet_name, self)
        if dlg.exec() == QDialog.Accepted:
            win = self.window()
            if hasattr(win, 'statusBar'):
                win.statusBar().showMessage(f"Import rules saved for {snippet_name} — guide compiled")

    # Default grid size (matches Google Sheets)
    DEFAULT_ROWS = 1000
    DEFAULT_COLS = 26

    def _load(self):
        ws = self.ws
        rows = max(ws.max_row or 1, self.DEFAULT_ROWS)
        cols = max(ws.max_column or 1, self.DEFAULT_COLS)
        self.setRowCount(rows)
        self.setColumnCount(cols)

        # Column headers
        self.setHorizontalHeaderLabels(
            [get_column_letter(c + 1) for c in range(cols)]
        )

        # Column widths
        for ci in range(cols):
            col_letter = get_column_letter(ci + 1)
            if col_letter in ws.column_dimensions:
                w = ws.column_dimensions[col_letter].width
                if w:
                    self.setColumnWidth(ci, xl_col_width_to_px(w))

        # Row heights
        for ri in range(rows):
            if (ri + 1) in ws.row_dimensions:
                h = ws.row_dimensions[ri + 1].height
                if h:
                    self.setRowHeight(ri, xl_row_height_to_px(h))

        # Cells
        for ri in range(rows):
            for ci in range(cols):
                cell = ws.cell(row=ri + 1, column=ci + 1)
                item = _xl_cell_to_item(cell)
                self.setItem(ri, ci, item)

        # Merged cells
        for rng in ws.merged_cells.ranges:
            r1, c1 = rng.min_row - 1, rng.min_col - 1
            r2, c2 = rng.max_row - 1, rng.max_col - 1
            self.merges.append((r1, c1, r2, c2))
            self.setSpan(r1, c1, r2 - r1 + 1, c2 - c1 + 1)

    # ── Write back ───────────────────────────────────────────────────────

    def sync_to_ws(self):
        """Push current table state back to the openpyxl worksheet."""
        ws = self.ws

        # Unmerge all existing
        for rng in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(rng))

        for ri in range(self.rowCount()):
            for ci in range(self.columnCount()):
                cell = ws.cell(row=ri + 1, column=ci + 1)
                item = self.item(ri, ci)
                _item_to_xl_cell(item, cell)

        # Re-merge
        for r1, c1, r2, c2 in self.merges:
            ws.merge_cells(
                start_row=r1 + 1,
                start_column=c1 + 1,
                end_row=r2 + 1,
                end_column=c2 + 1,
            )

        # Column widths
        for ci in range(self.columnCount()):
            col_letter = get_column_letter(ci + 1)
            ws.column_dimensions[col_letter].width = px_to_xl_col_width(
                self.columnWidth(ci)
            )

        # Row heights
        for ri in range(self.rowCount()):
            ws.row_dimensions[ri + 1].height = px_to_xl_row_height(
                self.rowHeight(ri)
            )


# ── Template Picker Dialog ───────────────────────────────────────────────────

class TemplatePicker(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("New from Template")
        self.setMinimumSize(500, 400)
        self.chosen_wb = None
        self._chosen_path = None

        layout = QVBoxLayout(self)

        # Section: Built-in Templates
        layout.addWidget(QLabel("Built-in Templates"))
        self.builtin_list = QListWidget()
        for name in BUILTIN_TEMPLATES:
            self.builtin_list.addItem(name)
        self.builtin_list.itemDoubleClicked.connect(self._use_builtin)
        layout.addWidget(self.builtin_list)

        # Section: User Templates
        _ensure_templates_dir()
        user_templates = sorted(TEMPLATES_DIR.glob("*.xlsx"))
        if user_templates:
            sep = QFrame()
            sep.setFrameShape(QFrame.HLine)
            layout.addWidget(sep)
            layout.addWidget(QLabel("My Templates"))
            self.user_list = QListWidget()
            for p in user_templates:
                item = QListWidgetItem(p.stem)
                item.setData(Qt.UserRole, str(p))
                self.user_list.addItem(item)
            self.user_list.itemDoubleClicked.connect(self._use_user)
            layout.addWidget(self.user_list)

            del_btn = QPushButton("Delete Selected Template")
            del_btn.clicked.connect(self._delete_user_template)
            layout.addWidget(del_btn)
        else:
            self.user_list = None

        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        use_btn = QPushButton("Use Template")
        use_btn.setDefault(True)
        use_btn.clicked.connect(self._use_selected)
        btn_layout.addWidget(use_btn)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

    def _use_selected(self):
        # Check builtin first, then user
        item = self.builtin_list.currentItem()
        if item and self.builtin_list.hasFocus():
            self._use_builtin(item)
            return
        if self.user_list:
            item = self.user_list.currentItem()
            if item:
                self._use_user(item)
                return
        # Fallback: use whatever is selected in builtin
        item = self.builtin_list.currentItem()
        if item:
            self._use_builtin(item)

    def _use_builtin(self, item):
        name = item.text()
        builder = BUILTIN_TEMPLATES.get(name)
        if builder:
            self.chosen_wb = builder()
            self.accept()

    def _use_user(self, item):
        path = item.data(Qt.UserRole)
        if path and Path(path).exists():
            try:
                self.chosen_wb = openpyxl.load_workbook(path)
                self._chosen_path = path
                self.accept()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to load template:\n{e}")

    def _delete_user_template(self):
        if not self.user_list:
            return
        item = self.user_list.currentItem()
        if not item:
            return
        path = item.data(Qt.UserRole)
        name = item.text()
        reply = QMessageBox.question(
            self, "Delete Template",
            f"Delete template \"{name}\"?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            try:
                Path(path).unlink()
                self.user_list.takeItem(self.user_list.row(item))
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete:\n{e}")


# ── Rules Editor Dialog ──────────────────────────────────────────────────────

RULE_TYPES = [
    "not_empty", "number_only", "max_length", "font_size_min",
    "font_size_max", "wrap_required", "fill_required", "row_height",
]

RULE_PARAMS = {
    "max_length": "max_length",
    "font_size_min": "min_size",
    "font_size_max": "max_size",
    "row_height": "height",
}


class RulesEditor(QDialog):
    def __init__(self, rules, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Rules")
        self.setMinimumSize(600, 450)
        self._rules = [dict(r) for r in rules]  # deep-ish copy
        for entry in self._rules:
            entry["rules"] = [dict(c) for c in entry.get("rules", [])]

        layout = QHBoxLayout(self)

        # Left: entry list
        left = QVBoxLayout()
        left.addWidget(QLabel("Rule Entries"))
        self.entry_list = QListWidget()
        self.entry_list.currentRowChanged.connect(self._show_entry)
        left.addWidget(self.entry_list)
        btn_row = QHBoxLayout()
        add_entry = QPushButton("Add Entry")
        add_entry.clicked.connect(self._add_entry)
        btn_row.addWidget(add_entry)
        del_entry = QPushButton("Remove Entry")
        del_entry.clicked.connect(self._del_entry)
        btn_row.addWidget(del_entry)
        left.addLayout(btn_row)
        layout.addLayout(left)

        # Right: entry detail
        right = QVBoxLayout()
        right.addWidget(QLabel("Range (e.g. A1:D10):"))
        self.range_edit = QLineEdit()
        self.range_edit.textChanged.connect(self._update_range)
        right.addWidget(self.range_edit)

        right.addWidget(QLabel("Conditions"))
        self.cond_list = QListWidget()
        right.addWidget(self.cond_list)

        cond_btn_row = QHBoxLayout()
        add_cond = QPushButton("Add Condition")
        add_cond.clicked.connect(self._add_cond)
        cond_btn_row.addWidget(add_cond)
        del_cond = QPushButton("Remove Condition")
        del_cond.clicked.connect(self._del_cond)
        cond_btn_row.addWidget(del_cond)
        right.addLayout(cond_btn_row)

        right.addStretch()

        ok_btn = QPushButton("OK")
        ok_btn.setDefault(True)
        ok_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        bottom = QHBoxLayout()
        bottom.addStretch()
        bottom.addWidget(ok_btn)
        bottom.addWidget(cancel_btn)
        right.addLayout(bottom)
        layout.addLayout(right)

        self._refresh_entries()

    def _refresh_entries(self):
        self.entry_list.clear()
        for entry in self._rules:
            self.entry_list.addItem(entry.get("range", "???"))

    def _show_entry(self, row):
        self.cond_list.clear()
        self.range_edit.blockSignals(True)
        if 0 <= row < len(self._rules):
            entry = self._rules[row]
            self.range_edit.setText(entry.get("range", ""))
            for cond in entry.get("rules", []):
                label = cond.get("type", "")
                param_key = RULE_PARAMS.get(label)
                if param_key and param_key in cond:
                    label += f" ({cond[param_key]})"
                self.cond_list.addItem(label)
        else:
            self.range_edit.setText("")
        self.range_edit.blockSignals(False)

    def _update_range(self, text):
        row = self.entry_list.currentRow()
        if 0 <= row < len(self._rules):
            self._rules[row]["range"] = text
            self.entry_list.currentItem().setText(text)

    def _add_entry(self):
        self._rules.append({"range": "A1:A1", "rules": []})
        self._refresh_entries()
        self.entry_list.setCurrentRow(len(self._rules) - 1)

    def _del_entry(self):
        row = self.entry_list.currentRow()
        if 0 <= row < len(self._rules):
            self._rules.pop(row)
            self._refresh_entries()

    def _add_cond(self):
        row = self.entry_list.currentRow()
        if row < 0 or row >= len(self._rules):
            return
        rtype, ok = QInputDialog.getItem(
            self, "Add Condition", "Rule type:", RULE_TYPES, 0, False,
        )
        if not ok:
            return
        cond = {"type": rtype}
        param_key = RULE_PARAMS.get(rtype)
        if param_key:
            val, ok2 = QInputDialog.getInt(self, "Parameter", f"{param_key}:", 10, 0, 99999)
            if not ok2:
                return
            cond[param_key] = val
        self._rules[row]["rules"].append(cond)
        self._show_entry(row)

    def _del_cond(self):
        entry_row = self.entry_list.currentRow()
        cond_row = self.cond_list.currentRow()
        if 0 <= entry_row < len(self._rules):
            conds = self._rules[entry_row].get("rules", [])
            if 0 <= cond_row < len(conds):
                conds.pop(cond_row)
                self._show_entry(entry_row)

    def get_rules(self):
        return self._rules


# ── Import Pipeline (OCR + Ollama) ───────────────────────────────────────────

OLLAMA_MODEL = "llama3.3:latest"
OLLAMA_URL = "http://localhost:11434/api/chat"

CALL_SHEET_SYSTEM_PROMPT = """You are a JSON data extractor. You ONLY output valid JSON. No explanations, no markdown, no commentary. Just a single JSON object matching EXACTLY the schema provided."""


def _snippet_guide_path(snippet_name):
    """Return the path to a snippet's per-snippet .guide.json file."""
    return SNIPPETS_DIR / f"{snippet_name}.guide.json"


def _load_snippet_rules_json(snippet_name):
    """Load a snippet's per-cell rules. Returns dict {cell_ref: rule_text, "_description": "..."}."""
    p = _snippet_guide_path(snippet_name)
    if p.exists():
        try:
            return json.loads(p.read_text())
        except Exception:
            pass
    # Migrate from old .guide.md if it exists
    old_md = SNIPPETS_DIR / f"{snippet_name}.guide.md"
    if old_md.exists():
        return _migrate_guide_md(snippet_name, old_md)
    return {}


def _migrate_guide_md(snippet_name, md_path):
    """Convert old .guide.md to .guide.json format."""
    text = md_path.read_text()
    rules = {"_description": ""}
    desc_lines = []
    for line in text.strip().splitlines():
        line_s = line.strip()
        if line_s.startswith("- ") and ":" in line_s[2:6]:
            # Parse "- A1: description"
            rest = line_s[2:]
            ref, _, desc = rest.partition(":")
            ref = ref.strip()
            rules[ref] = desc.strip()
        else:
            desc_lines.append(line)
    rules["_description"] = "\n".join(desc_lines).strip()
    # Save as JSON and remove old md
    _save_snippet_rules_json(snippet_name, rules)
    md_path.unlink()
    return rules


def _save_snippet_rules_json(snippet_name, rules):
    """Save a snippet's per-cell rules as JSON."""
    _ensure_snippets_dir()
    p = _snippet_guide_path(snippet_name)
    # Remove empty entries
    clean = {k: v for k, v in rules.items() if v}
    if clean:
        p.write_text(json.dumps(clean, indent=2))
    elif p.exists():
        p.unlink()


def _snippet_cell_refs(snippet_name):
    """Read a snippet xlsx and return list of (cell_ref, value) for non-empty cells."""
    path = SNIPPETS_DIR / f"{snippet_name}.xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    cells = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for c in row:
            if c.value is not None:
                ref = f"{get_column_letter(c.column)}{c.row}"
                cells.append((ref, str(c.value)))
    return cells


def _compile_import_guide():
    """Compile all per-snippet .guide.json files into the master import_guide.md."""
    snippets = list_snippets()
    sections = []
    for name in snippets:
        rules = _load_snippet_rules_json(name)
        if not rules:
            continue
        # Only include snippets that have at least one cell rule
        cell_rules = {k: v for k, v in rules.items() if k != "_description" and v}
        if not cell_rules:
            continue
        cell_refs = _snippet_cell_refs(name)
        lines = [f"## {name}", ""]
        desc = rules.get("_description", "")
        if desc:
            lines.append(desc)
            lines.append("")
        # Build cell descriptions from rules
        lines.append("Fill these cells:")
        for ref, rule_text in cell_rules.items():
            lines.append(f"- {ref}: {rule_text}")
        lines.append("")
        # Build JSON template from all non-empty cells in snippet
        if cell_refs:
            lines.append("Return JSON:")
            lines.append("```json")
            lines.append("{")
            json_lines = [f'  "{ref}": ""' for ref, _ in cell_refs]
            lines.append(",\n".join(json_lines))
            lines.append("}")
            lines.append("```")
            lines.append("")
        lines.append('IMPORTANT: If a field\'s data is NOT found in the source text, you MUST return "" for that field. Do NOT guess or use placeholder values. Only fill cells with data that is explicitly present in the source text.')
        lines.append("")
        sections.append("\n".join(lines))

    header = "# Import Guide\n\nEach snippet section below tells the AI what data to extract and which cells to fill.\nWhen importing, only the active snippet's section is sent to the AI.\n\n---\n"
    full = header + "\n".join(sections)
    IMPORT_GUIDE_PATH.parent.mkdir(parents=True, exist_ok=True)
    IMPORT_GUIDE_PATH.write_text(full)
    return full


class SnippetRulesEditor(QDialog):
    """Visual editor for all snippets: list on left, spreadsheet in middle, rule editor on right."""

    def __init__(self, snippet_name=None, parent=None):
        super().__init__(parent)
        self._snippet_name = None
        self._rules = {}
        self._current_ref = None
        self.setWindowTitle("Snippet Rules Editor")
        self.resize(1100, 600)

        main_layout = QVBoxLayout(self)

        # Top: description row (updates per snippet)
        desc_row = QHBoxLayout()
        desc_row.addWidget(QLabel("Description:"))
        self._desc_edit = QLineEdit()
        self._desc_edit.setPlaceholderText("Brief description of this snippet (optional)")
        desc_row.addWidget(self._desc_edit)
        main_layout.addLayout(desc_row)

        # Three-panel splitter
        splitter = QSplitter(Qt.Horizontal)

        # Left panel: snippet list
        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(0, 0, 0, 0)
        list_label = QLabel("Snippets")
        list_label.setStyleSheet("font-weight: bold; font-size: 13px;")
        left_layout.addWidget(list_label)
        self._snippet_list = QListWidget()
        self._snippet_list.currentRowChanged.connect(self._snippet_selected)
        left_layout.addWidget(self._snippet_list)
        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(0, 0, 0, 0)
        btn_row.setSpacing(2)
        add_btn = QPushButton("+")
        add_btn.setToolTip("Add snippet from selection")
        add_btn.setFixedWidth(30)
        add_btn.clicked.connect(self._add_snippet)
        btn_row.addWidget(add_btn)
        rename_btn = QPushButton("Rename")
        rename_btn.clicked.connect(self._rename_snippet)
        btn_row.addWidget(rename_btn)
        del_btn = QPushButton("−")
        del_btn.setToolTip("Delete snippet")
        del_btn.setFixedWidth(30)
        del_btn.clicked.connect(self._delete_snippet)
        btn_row.addWidget(del_btn)
        left_layout.addLayout(btn_row)
        splitter.addWidget(left)

        # Middle panel: spreadsheet preview
        self._table = QTableWidget()
        self._table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._table.setSelectionMode(QTableWidget.SingleSelection)
        self._table.setItemDelegate(BorderDelegate(self._table))
        self._table.currentCellChanged.connect(self._cell_selected)
        splitter.addWidget(self._table)

        # Right panel: cell rule editor
        right = QWidget()
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(8, 0, 0, 0)

        self._ref_label = QLabel("Select a cell")
        self._ref_label.setStyleSheet("font-size: 16px; font-weight: bold;")
        right_layout.addWidget(self._ref_label)

        self._val_label = QLabel("")
        self._val_label.setStyleSheet("color: #5F6368; font-size: 12px;")
        self._val_label.setWordWrap(True)
        right_layout.addWidget(self._val_label)

        rule_label = QLabel("Rule for this cell:")
        right_layout.addWidget(rule_label)

        self._rule_edit = QPlainTextEdit()
        self._rule_edit.setFont(QFont("Menlo", 12))
        self._rule_edit.setPlaceholderText(
            "Describe what data goes here.\n"
            'e.g. "Production title, formatted as TITLE - CALL SHEET - X Pages"'
        )
        self._rule_edit.textChanged.connect(self._rule_text_changed)
        right_layout.addWidget(self._rule_edit)

        self._status_label = QLabel("")
        self._status_label.setStyleSheet("color: #5F6368; font-size: 11px;")
        self._status_label.setWordWrap(True)
        right_layout.addWidget(self._status_label)

        splitter.addWidget(right)
        splitter.setSizes([160, 500, 340])
        main_layout.addWidget(splitter)

        # Buttons
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        save_btn = QPushButton("Save && Compile")
        save_btn.setStyleSheet(
            "QPushButton { background-color: #1A73E8; color: white; "
            "font-weight: bold; padding: 6px 20px; border-radius: 4px; }"
            "QPushButton:hover { background-color: #1557B0; }"
        )
        save_btn.clicked.connect(self._save_all)
        btn_row.addWidget(save_btn)
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.reject)
        btn_row.addWidget(close_btn)
        main_layout.addLayout(btn_row)

        # Populate snippet list
        self._all_rules = {}  # {snippet_name: rules_dict}
        self._populate_snippets()

        # Select the requested snippet, or the first one
        if snippet_name:
            for i in range(self._snippet_list.count()):
                if self._snippet_list.item(i).data(Qt.UserRole) == snippet_name:
                    self._snippet_list.setCurrentRow(i)
                    break
        elif self._snippet_list.count() > 0:
            self._snippet_list.setCurrentRow(0)

    def _populate_snippets(self):
        """Fill the snippet list with all available snippets."""
        snippets = list_snippets()
        for name in snippets:
            rules = _load_snippet_rules_json(name)
            self._all_rules[name] = rules
            item = QListWidgetItem()
            item.setData(Qt.UserRole, name)
            item.setText(name)
            self._snippet_list.addItem(item)

    def _snippet_selected(self, row):
        """Switch to a different snippet."""
        # Save current snippet's description before switching
        self._commit_current()

        if row < 0 or row >= self._snippet_list.count():
            return
        name = self._snippet_list.item(row).data(Qt.UserRole)
        self._snippet_name = name
        self._rules = self._all_rules.get(name, {})
        self._current_ref = None

        # Update description
        self._desc_edit.setText(self._rules.get("_description", ""))

        # Reset right panel
        self._ref_label.setText("Select a cell")
        self._val_label.setText("")
        self._rule_edit.blockSignals(True)
        self._rule_edit.setPlainText("")
        self._rule_edit.blockSignals(False)

        # Load spreadsheet preview
        self._load_snippet_preview()
        self._update_status()

    def _commit_current(self):
        """Save in-memory state for the current snippet before switching."""
        if self._snippet_name and self._snippet_name in self._all_rules:
            self._all_rules[self._snippet_name]["_description"] = self._desc_edit.text().strip()

    def _load_snippet_preview(self):
        """Load the snippet xlsx into the mini spreadsheet."""
        self._table.clear()
        self._table.setRowCount(0)
        self._table.setColumnCount(0)
        if not self._snippet_name:
            return
        path = SNIPPETS_DIR / f"{self._snippet_name}.xlsx"
        if not path.exists():
            return
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        rows = ws.max_row or 1
        cols = ws.max_column or 1
        self._table.setRowCount(rows)
        self._table.setColumnCount(cols)
        self._table.setHorizontalHeaderLabels(
            [get_column_letter(c + 1) for c in range(cols)]
        )
        for row in ws.iter_rows(min_row=1, max_row=rows, max_col=cols):
            for cell in row:
                r, c = cell.row - 1, cell.column - 1
                item = _xl_cell_to_item(cell)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                ref = f"{get_column_letter(cell.column)}{cell.row}"
                if ref in self._rules and self._rules[ref]:
                    item.setData(RULE_ROLE, True)
                self._table.setItem(r, c, item)
        for ci in range(cols):
            col_letter = get_column_letter(ci + 1)
            if col_letter in ws.column_dimensions:
                w = ws.column_dimensions[col_letter].width
                if w:
                    self._table.setColumnWidth(ci, xl_col_width_to_px(w))
        for ri in range(rows):
            if (ri + 1) in ws.row_dimensions:
                h = ws.row_dimensions[ri + 1].height
                if h:
                    self._table.setRowHeight(ri, int(h * 1.33))
        for mg in ws.merged_cells.ranges:
            r1, c1, r2, c2 = mg.min_row - 1, mg.min_col - 1, mg.max_row - 1, mg.max_col - 1
            self._table.setSpan(r1, c1, r2 - r1 + 1, c2 - c1 + 1)

    def _cell_selected(self, row, col, prev_row, prev_col):
        """Update right panel when a cell is clicked."""
        if row < 0 or col < 0:
            return
        ref = f"{get_column_letter(col + 1)}{row + 1}"
        self._current_ref = ref
        self._ref_label.setText(ref)
        item = self._table.item(row, col)
        val = item.text() if item else ""
        self._val_label.setText(f"Current value: {val}" if val else "(empty cell)")
        self._rule_edit.blockSignals(True)
        self._rule_edit.setPlainText(self._rules.get(ref, ""))
        self._rule_edit.blockSignals(False)

    def _rule_text_changed(self):
        """Save rule text back to dict as user types."""
        if not self._current_ref or not self._snippet_name:
            return
        self._rules[self._current_ref] = self._rule_edit.toPlainText()
        # Update cell highlight
        ref = self._current_ref
        try:
            col_idx = column_index_from_string(ref.rstrip("0123456789")) - 1
            row_idx = int(ref.lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ")) - 1
            item = self._table.item(row_idx, col_idx)
            if item:
                if self._rule_edit.toPlainText().strip():
                    item.setBackground(QBrush(QColor("#E8F0FE")))
                else:
                    item.setBackground(QBrush(QColor("#FFFFFF")))
        except Exception:
            pass
        self._update_status()
        self._update_list_label()

    def _update_status(self):
        count = sum(1 for k, v in self._rules.items() if k != "_description" and v)
        self._status_label.setText(f"{count} cell(s) have rules defined")

    def _update_list_label(self):
        """Update the current snippet's list item text."""
        row = self._snippet_list.currentRow()
        if row < 0:
            return
        item = self._snippet_list.item(row)
        name = item.data(Qt.UserRole)
        item.setText(name)

    def _add_snippet(self):
        """Save the current selection from the main sheet as a new snippet."""
        parent_win = self.parent()
        if not parent_win or not hasattr(parent_win, '_selected_range'):
            QMessageBox.information(self, "Add Snippet",
                "Close this editor, select cells in the spreadsheet, then use 'Save Selection as Snippet'.")
            return
        rng = parent_win._selected_range()
        if rng is None:
            QMessageBox.information(self, "Add Snippet",
                "No cells selected in the main spreadsheet.\n"
                "Close this editor, select cells, then reopen.")
            return
        r1, c1, r2, c2 = rng
        name, ok = QInputDialog.getText(self, "Add Snippet", "Snippet name:")
        if not ok or not name.strip():
            return
        name = name.strip()
        if (SNIPPETS_DIR / f"{name}.xlsx").exists():
            QMessageBox.warning(self, "Add Snippet", f'A snippet named "{name}" already exists.')
            return
        sv = parent_win._sheet()
        if sv is None:
            return
        save_snippet(name, sv, r1, c1, r2, c2)
        # Add to internal state and list
        self._all_rules[name] = {}
        item = QListWidgetItem()
        item.setData(Qt.UserRole, name)
        item.setText(name)
        self._snippet_list.addItem(item)
        self._snippet_list.setCurrentItem(item)

    def _delete_snippet(self):
        """Delete the currently selected snippet and all its sidecar files."""
        row = self._snippet_list.currentRow()
        if row < 0:
            return
        item = self._snippet_list.item(row)
        name = item.data(Qt.UserRole)
        reply = QMessageBox.question(
            self, "Delete Snippet",
            f'Delete snippet "{name}" and all its rules? This cannot be undone.',
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        # Delete all sidecar files
        for ext in [".xlsx", ".meta.json", ".guide.json", ".guide.md"]:
            p = SNIPPETS_DIR / f"{name}{ext}"
            if p.exists():
                p.unlink()
        # Remove from internal state
        self._all_rules.pop(name, None)
        if self._snippet_name == name:
            self._snippet_name = None
            self._current_ref = None
        self._snippet_list.takeItem(row)

    def _rename_snippet(self):
        """Rename the currently selected snippet and all its sidecar files."""
        row = self._snippet_list.currentRow()
        if row < 0:
            return
        item = self._snippet_list.item(row)
        old_name = item.data(Qt.UserRole)
        new_name, ok = QInputDialog.getText(
            self, "Rename Snippet", "New name:", text=old_name
        )
        if not ok or not new_name.strip() or new_name.strip() == old_name:
            return
        new_name = new_name.strip()
        # Check for conflict
        if (SNIPPETS_DIR / f"{new_name}.xlsx").exists():
            QMessageBox.warning(self, "Rename", f'A snippet named "{new_name}" already exists.')
            return
        # Rename all sidecar files
        for ext in [".xlsx", ".meta.json", ".guide.json", ".guide.md"]:
            old_path = SNIPPETS_DIR / f"{old_name}{ext}"
            if old_path.exists():
                old_path.rename(SNIPPETS_DIR / f"{new_name}{ext}")
        # Update internal state
        rules = self._all_rules.pop(old_name, {})
        self._all_rules[new_name] = rules
        if self._snippet_name == old_name:
            self._snippet_name = new_name
        # Update list item
        item.setData(Qt.UserRole, new_name)
        self._update_list_label()

    def _save_all(self):
        """Save all snippet rules and compile the import guide."""
        self._commit_current()
        for name, rules in self._all_rules.items():
            _save_snippet_rules_json(name, rules)
        _compile_import_guide()
        self.accept()


def _get_snippet_guide(snippet_name):
    """Extract the guide section for a specific snippet from the import guide."""
    if not IMPORT_GUIDE_PATH.exists():
        return None
    guide = IMPORT_GUIDE_PATH.read_text()
    # Find the section for this snippet
    marker = f"## {snippet_name}"
    start = guide.find(marker)
    if start < 0:
        return None
    # Find the next ## section or end of file
    next_section = guide.find("\n## ", start + len(marker))
    if next_section < 0:
        section = guide[start:]
    else:
        section = guide[start:next_section]
    return section.strip()


def _build_snippet_prompt(snippet_name):
    """Build AI prompt for filling a specific snippet."""
    guide_section = _get_snippet_guide(snippet_name)
    if not guide_section:
        return None

    return f"""You are extracting data to fill a spreadsheet snippet.

{guide_section}

CRITICAL RULES:
- Return ONLY the JSON object with cell references as keys.
- If a field is NOT explicitly found in the source text, you MUST use "" (empty string).
- Do NOT guess, infer, or use placeholder values. Only use data that appears in the text.
- No explanation, no commentary.

Text:
"""


def _ocr_image(path):
    """Extract text from an image using macOS Vision framework."""
    try:
        import Vision
        from Cocoa import NSURL
        from Quartz import CIImage

        url = NSURL.fileURLWithPath_(str(path))
        request = Vision.VNRecognizeTextRequest.alloc().init()
        request.setRecognitionLevel_(Vision.VNRequestTextRecognitionLevelAccurate)
        handler = Vision.VNImageRequestHandler.alloc().initWithURL_options_(url, {})
        handler.performRequests_error_([request], None)

        results = request.results()
        if not results:
            return ""
        lines = []
        for obs in results:
            text = obs.topCandidates_(1)[0].string()
            lines.append(text)
        return "\n".join(lines)
    except Exception as e:
        return f"[OCR Error: {e}]"


def _extract_text_from_pdf(path):
    """Extract text from a PDF using macOS Vision OCR."""
    try:
        import Vision
        from Cocoa import NSURL
        from Quartz import (
            PDFDocument, CGPDFDocumentCreateWithURL,
        )
        import Quartz

        url = NSURL.fileURLWithPath_(str(path))
        pdf_doc = PDFDocument.alloc().initWithURL_(url)
        if not pdf_doc:
            return "[Could not open PDF]"

        all_text = []
        for i in range(pdf_doc.pageCount()):
            page = pdf_doc.pageAtIndex_(i)
            page_text = page.string()
            if page_text:
                all_text.append(page_text)
            else:
                # Fall back to OCR for scanned pages
                # Get page as image and OCR it
                pass
        return "\n".join(all_text) if all_text else "[No text found in PDF]"
    except Exception as e:
        return f"[PDF Error: {e}]"


def _extract_text_from_xlsx(path):
    """Extract text from an xlsx file (first/active sheet only, skip empty rows)."""
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        lines = [f"[Sheet: {ws.title}]"]
        for row in ws.iter_rows(values_only=True):
            vals = [str(v) for v in row if v is not None]
            if vals:
                lines.append("\t".join(vals))
        return "\n".join(lines)
    except Exception as e:
        return f"[XLSX Error: {e}]"


def _extract_text_from_file(path):
    """Route file to appropriate text extractor."""
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix in (".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".heic", ".webp"):
        return _ocr_image(str(p))
    elif suffix == ".pdf":
        return _extract_text_from_pdf(str(p))
    elif suffix == ".xlsx":
        return _extract_text_from_xlsx(str(p))
    elif suffix in (".txt", ".csv"):
        return p.read_text(errors="replace")
    else:
        return f"[Unsupported file type: {suffix}]"


def _query_ollama(system_prompt, user_prompt, progress_fn=None):
    """Send a chat to Ollama with streaming and return the response text."""
    import requests
    import time
    try:
        resp = requests.post(OLLAMA_URL, json={
            "model": OLLAMA_MODEL,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            "format": "json",
            "stream": True,
        }, timeout=300, stream=True)
        resp.raise_for_status()

        full_response = ""
        token_count = 0
        start_time = time.time()

        for line in resp.iter_lines():
            if not line:
                continue
            chunk = json.loads(line)
            content = chunk.get("message", {}).get("content", "")
            if content:
                full_response += content
                token_count += 1
                if progress_fn and token_count % 10 == 0:
                    elapsed = int(time.time() - start_time)
                    progress_fn(f"Generating... {token_count} tokens, {elapsed}s elapsed")
            if chunk.get("done"):
                break

        if progress_fn:
            elapsed = int(time.time() - start_time)
            progress_fn(f"Done — {token_count} tokens in {elapsed}s")

        return full_response
    except Exception as e:
        return f"[Ollama Error: {e}]"


def _parse_call_sheet_data(raw_text, log_fn=None):
    """Send extracted text to Ollama to parse into structured JSON."""
    user_msg = _build_import_prompt() + raw_text
    response = _query_ollama(CALL_SHEET_SYSTEM_PROMPT, user_msg, progress_fn=log_fn)

    if response.startswith("[Ollama"):
        if log_fn:
            log_fn(response)
        return None

    if log_fn:
        # Show first 300 chars of response for debugging
        preview = response[:300].replace("\n", " ")
        log_fn(f"AI response preview: {preview}...")

    # Strip <think>...</think> blocks (deepseek-r1)
    import re
    cleaned = re.sub(r"<think>.*?</think>", "", response, flags=re.DOTALL).strip()

    # Strip markdown code fences
    if "```" in cleaned:
        lines = cleaned.split("\n")
        lines = [l for l in lines if not l.strip().startswith("```")]
        cleaned = "\n".join(lines)

    # Find JSON object — look for outermost { }
    start = cleaned.find("{")
    end = cleaned.rfind("}") + 1
    if start >= 0 and end > start:
        json_str = cleaned[start:end]
        try:
            return json.loads(json_str)
        except json.JSONDecodeError as e:
            if log_fn:
                log_fn(f"JSON parse error: {e}")
                log_fn(f"JSON snippet: {json_str[:200]}...")

    if log_fn:
        log_fn(f"Full response length: {len(response)} chars")
    return None


def _fill_snippet_cells(wb, cell_data, sheet_index=0):
    """Fill cells in a workbook using cell-reference keys like {"A1": "value", "G2": "value"}.

    Returns the workbook.
    """
    ws = wb.worksheets[sheet_index] if sheet_index < len(wb.worksheets) else wb.active
    for ref, value in cell_data.items():
        if value:  # skip empty strings
            try:
                ws[ref].value = value
            except (ValueError, KeyError):
                pass
    return wb


def _import_to_snippet(snippet_name, source_text, log_fn=None):
    """Extract data from source text and fill a snippet. Returns a workbook."""
    prompt = _build_snippet_prompt(snippet_name)
    if not prompt:
        if log_fn:
            log_fn(f"No import guide section found for '{snippet_name}'")
        return None

    if log_fn:
        log_fn(f"Extracting data for: {snippet_name}")

    response = _query_ollama(CALL_SHEET_SYSTEM_PROMPT, prompt + source_text,
                             progress_fn=log_fn)

    if response.startswith("[Ollama"):
        if log_fn:
            log_fn(response)
        return None

    # Strip think blocks
    import re
    cleaned = re.sub(r"<think>.*?</think>", "", response, flags=re.DOTALL).strip()

    # Parse JSON
    start = cleaned.find("{")
    end = cleaned.rfind("}") + 1
    if start < 0 or end <= start:
        if log_fn:
            log_fn(f"No JSON found in response ({len(response)} chars)")
        return None

    try:
        cell_data = json.loads(cleaned[start:end])
    except json.JSONDecodeError as e:
        if log_fn:
            log_fn(f"JSON parse error: {e}")
        return None

    if log_fn:
        log_fn(f"Got {len(cell_data)} cell values")
        for ref, val in cell_data.items():
            if val:
                log_fn(f"  {ref} = {val}")

    # Load the snippet as a workbook and fill
    path = SNIPPETS_DIR / f"{snippet_name}.xlsx"
    if not path.exists():
        if log_fn:
            log_fn(f"Snippet file not found: {path}")
        return None

    wb = openpyxl.load_workbook(str(path))
    return wb, cell_data


def _parse_guide_labels(snippet_name):
    """Extract cell ref -> label mapping from the import guide."""
    section = _get_snippet_guide(snippet_name)
    if not section:
        return {}
    labels = {}
    for line in section.split("\n"):
        line = line.strip()
        if line.startswith("- ") and ":" in line:
            # Parse "- A1: Production title, formatted as ..."
            parts = line[2:].split(":", 1)
            ref = parts[0].strip()
            desc = parts[1].strip() if len(parts) > 1 else ref
            # Truncate at first parenthetical example
            if " (e.g." in desc:
                desc = desc[:desc.index(" (e.g.")].strip()
            if " (e.g " in desc:
                desc = desc[:desc.index(" (e.g ")].strip()
            labels[ref] = desc
    return labels


def _prompt_missing_fields(snippet_name, cell_data, parent=None):
    """Prompt the user one at a time for any empty fields. Returns updated cell_data."""
    labels = _parse_guide_labels(snippet_name)
    result = dict(cell_data)

    # Find empty fields from the guide
    missing = []
    for ref in labels:
        if not result.get(ref):
            missing.append(ref)

    if not missing:
        return result

    for i, ref in enumerate(missing):
        label = labels.get(ref, ref)
        text, ok = QInputDialog.getText(
            parent,
            f"Missing field ({i + 1}/{len(missing)})",
            f"{ref} — {label}:",
        )
        if not ok:
            # User cancelled — stop prompting, keep what we have
            break
        if text:
            result[ref] = text

    return result


def _fill_call_sheet(data):
    """Build a call sheet from snippets and fill with structured data.

    Uses snippet meta files for dynamic row expansion. Returns a workbook.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Call Sheet"

    # Snippet order and their data key + count mapping
    snippet_data_map = {
        "Call Sheet_Header": "header",
        "Call Sheet_Scenes": "scenes",
        "Call Sheet_Cast": "cast",
        "Call Sheet_Crew": "crew",
        "Call Sheet_KeyPersonel_Weather_Safety": "key_personnel",
        "Call Sheet_Shooting Order_Company Moves": "shooting_order",
        "Call Sheet_Notes": "notes",
    }
    snippet_order = [
        "Call Sheet_Header",
        "Call Sheet_Scenes",
        "Call Sheet_Cast",
        "Call Sheet_Crew",
        "Call Sheet_KeyPersonel_Weather_Safety",
        "Call Sheet_Shooting Order_Company Moves",
        "Call Sheet_Notes",
    ]

    # Determine data counts for expandable snippets
    data_counts = {}
    if data.get("scenes"):
        data_counts["Call Sheet_Scenes"] = len(data["scenes"])
    if data.get("cast"):
        data_counts["Call Sheet_Cast"] = len(data["cast"])
    if data.get("crew"):
        # Crew uses two-column layout, so half the count (rounded up)
        data_counts["Call Sheet_Crew"] = (len(data["crew"]) + 1) // 2

    # Build snippets
    current_row = 0
    snippet_positions = {}  # track where each snippet starts for data filling
    for name in snippet_order:
        path = SNIPPETS_DIR / f"{name}.xlsx"
        if not path.exists():
            continue
        snippet_positions[name] = current_row
        dc = data_counts.get(name)
        rows_written = _build_snippet_to_ws(ws, name, current_row, data_count=dc)
        current_row += rows_written

    # Now fill in the actual data values
    header = data.get("header", {})
    if header and "Call Sheet_Header" in snippet_positions:
        base = snippet_positions["Call Sheet_Header"]
        if header.get("title"):
            ws.cell(row=base + 1, column=1).value = header["title"]
        if header.get("shoot_day"):
            ws.cell(row=base + 2, column=1).value = header["shoot_day"]
        if header.get("date"):
            ws.cell(row=base + 2, column=7).value = header["date"]
        if header.get("general_call"):
            ws.cell(row=base + 3, column=4).value = header["general_call"]
        if header.get("first_location"):
            ws.cell(row=base + 3, column=10).value = header["first_location"]
        if header.get("lunch"):
            ws.cell(row=base + 4, column=4).value = header["lunch"]
        if header.get("wrap_target"):
            ws.cell(row=base + 4, column=10).value = header["wrap_target"]

    scenes = data.get("scenes", [])
    if scenes and "Call Sheet_Scenes" in snippet_positions:
        base = snippet_positions["Call Sheet_Scenes"]
        meta = _load_snippet_meta("Call Sheet_Scenes")
        data_start = base + len(meta["header_rows"]) + 1 if meta else base + 3
        for i, sc in enumerate(scenes):
            r = data_start + i
            ws.cell(row=r, column=1).value = sc.get("scene_num", "")
            ws.cell(row=r, column=2).value = sc.get("description", "")
            ws.cell(row=r, column=5).value = sc.get("int_ext", "")
            ws.cell(row=r, column=6).value = sc.get("day_night", "")
            ws.cell(row=r, column=7).value = sc.get("pages", "")
            ws.cell(row=r, column=8).value = sc.get("cast", "")
            ws.cell(row=r, column=9).value = sc.get("location_notes", "")

    cast = data.get("cast", [])
    if cast and "Call Sheet_Cast" in snippet_positions:
        base = snippet_positions["Call Sheet_Cast"]
        meta = _load_snippet_meta("Call Sheet_Cast")
        data_start = base + len(meta["header_rows"]) + 1 if meta else base + 3
        for i, c in enumerate(cast):
            r = data_start + i
            ws.cell(row=r, column=1).value = c.get("number", str(i + 1))
            ws.cell(row=r, column=2).value = c.get("actor", "")
            ws.cell(row=r, column=3).value = c.get("character", "")
            ws.cell(row=r, column=4).value = c.get("status", "")
            ws.cell(row=r, column=5).value = c.get("set_call", "")
            ws.cell(row=r, column=6).value = c.get("hmu", "")
            ws.cell(row=r, column=7).value = c.get("phone", "")
            ws.cell(row=r, column=8).value = c.get("notes", "")

    crew = data.get("crew", [])
    if crew and "Call Sheet_Crew" in snippet_positions:
        base = snippet_positions["Call Sheet_Crew"]
        meta = _load_snippet_meta("Call Sheet_Crew")
        data_start = base + len(meta["header_rows"]) + 1 if meta else base + 3
        # Two-column layout: left side cols A-F, right side cols H-M
        for i, cr in enumerate(crew):
            row_offset = i // 2
            r = data_start + row_offset
            if i % 2 == 0:  # left side
                ws.cell(row=r, column=1).value = cr.get("position", "")
                ws.cell(row=r, column=3).value = cr.get("name", "")
                ws.cell(row=r, column=5).value = cr.get("phone", "")
                ws.cell(row=r, column=6).value = cr.get("call_time", "")
            else:  # right side
                ws.cell(row=r, column=8).value = cr.get("position", "")
                ws.cell(row=r, column=10).value = cr.get("name", "")
                ws.cell(row=r, column=12).value = cr.get("phone", "")
                ws.cell(row=r, column=13).value = cr.get("call_time", "")

    kp = data.get("key_personnel", {})
    if kp and "Call Sheet_KeyPersonel_Weather_Safety" in snippet_positions:
        base = snippet_positions["Call Sheet_KeyPersonel_Weather_Safety"]
        personnel = [
            ("director", 2), ("producer1", 3), ("producer2", 4),
            ("first_ad", 5), ("second_ad", 6), ("loc_mgr", 7),
        ]
        for key, row_off in personnel:
            person = kp.get(key, {})
            if person.get("name"):
                ws.cell(row=base + row_off, column=2).value = person["name"]
            if person.get("phone"):
                ws.cell(row=base + row_off, column=5).value = person["phone"]

    weather = data.get("weather", {})
    if weather and "Call Sheet_KeyPersonel_Weather_Safety" in snippet_positions:
        base = snippet_positions["Call Sheet_KeyPersonel_Weather_Safety"]
        if weather.get("sunrise"):
            ws.cell(row=base + 2, column=8).value = weather["sunrise"]
        if weather.get("sunset"):
            ws.cell(row=base + 3, column=8).value = weather["sunset"]
        if weather.get("high"):
            ws.cell(row=base + 4, column=8).value = weather["high"]
        if weather.get("low"):
            ws.cell(row=base + 5, column=8).value = weather["low"]
        if weather.get("precip"):
            ws.cell(row=base + 6, column=8).value = weather["precip"]
        if weather.get("condition"):
            ws.cell(row=base + 7, column=8).value = weather["condition"]
        if weather.get("hospital_name"):
            ws.cell(row=base + 3, column=10).value = weather["hospital_name"]
        if weather.get("hospital_address"):
            ws.cell(row=base + 5, column=10).value = weather["hospital_address"]
        if weather.get("hospital_phone"):
            ws.cell(row=base + 7, column=10).value = weather["hospital_phone"]
        if weather.get("alert"):
            ws.cell(row=base + 9, column=1).value = weather["alert"]

    shooting = data.get("shooting_order", "")
    if shooting and "Call Sheet_Shooting Order_Company Moves" in snippet_positions:
        base = snippet_positions["Call Sheet_Shooting Order_Company Moves"]
        ws.cell(row=base + 2, column=1).value = shooting

    notes = data.get("notes", {})
    if notes and "Call Sheet_Notes" in snippet_positions:
        base = snippet_positions["Call Sheet_Notes"]
        if notes.get("props"):
            ws.cell(row=base + 3, column=1).value = notes["props"]
        if notes.get("wardrobe"):
            ws.cell(row=base + 3, column=4).value = notes["wardrobe"]
        if notes.get("vehicles"):
            ws.cell(row=base + 3, column=7).value = notes["vehicles"]
        if notes.get("bg_atmosphere"):
            ws.cell(row=base + 3, column=10).value = notes["bg_atmosphere"]

    return wb


# ── Import Progress Dialog ──────────────────────────────────────────────────

SUPPORTED_IMPORT_EXTS = {".xlsx", ".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".heic", ".webp", ".bmp", ".csv", ".txt"}


class DropZoneLabel(QLabel):
    """A label that accepts drag-and-drop files."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setAlignment(Qt.AlignCenter)
        self.setText("Drop files here\n\nImages, PDFs, Spreadsheets\n\nor click Add Files below")
        self.setMinimumHeight(120)
        self.setStyleSheet(
            "QLabel { border: 2px dashed #BFBFBF; border-radius: 8px; "
            "color: #5F6368; font-size: 14px; padding: 20px; background: #F8F9FA; }"
        )
        self.files_dropped = None  # callback

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.setStyleSheet(
                "QLabel { border: 2px dashed #1A73E8; border-radius: 8px; "
                "color: #1A73E8; font-size: 14px; padding: 20px; background: #E8F0FE; }"
            )

    def dragLeaveEvent(self, event):
        self.setStyleSheet(
            "QLabel { border: 2px dashed #BFBFBF; border-radius: 8px; "
            "color: #5F6368; font-size: 14px; padding: 20px; background: #F8F9FA; }"
        )

    def dropEvent(self, event):
        self.setStyleSheet(
            "QLabel { border: 2px dashed #BFBFBF; border-radius: 8px; "
            "color: #5F6368; font-size: 14px; padding: 20px; background: #F8F9FA; }"
        )
        files = []
        for url in event.mimeData().urls():
            path = url.toLocalFile()
            if path and Path(path).suffix.lower() in SUPPORTED_IMPORT_EXTS:
                files.append(path)
        if files and self.files_dropped:
            self.files_dropped(files)


class ImportCallSheetDialog(QDialog):
    """Dialog that imports files, extracts text via AI, and fills a snippet."""

    def __init__(self, files=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Import Data to Snippet")
        self.setMinimumSize(550, 450)
        self.files = list(files) if files else []
        self.result_wb = None
        self._processing = False

        layout = QVBoxLayout(self)

        # Snippet selector
        snippet_row = QHBoxLayout()
        snippet_row.addWidget(QLabel("Fill snippet:"))
        self.snippet_combo = QComboBox()
        # Only show snippets that have a guide section
        for name in list_snippets():
            if _get_snippet_guide(name):
                self.snippet_combo.addItem(name)
        if self.snippet_combo.count() == 0:
            # Show all snippets if no guide sections exist yet
            for name in list_snippets():
                self.snippet_combo.addItem(name)
        snippet_row.addWidget(self.snippet_combo, 1)
        layout.addLayout(snippet_row)

        # Drop zone
        self.drop_zone = DropZoneLabel()
        self.drop_zone.files_dropped = self._add_files
        layout.addWidget(self.drop_zone)

        # File list
        self.file_list = QListWidget()
        layout.addWidget(self.file_list)

        # Buttons row
        btn_row = QHBoxLayout()
        add_btn = QPushButton("Add Files...")
        add_btn.clicked.connect(self._browse_files)
        btn_row.addWidget(add_btn)

        self.import_btn = QPushButton("Import")
        self.import_btn.setStyleSheet(
            "QPushButton { background-color: #1A73E8; color: white; "
            "font-weight: bold; padding: 6px 20px; border-radius: 4px; }"
            "QPushButton:hover { background-color: #1557B0; }"
        )
        self.import_btn.clicked.connect(self._start_import)
        btn_row.addWidget(self.import_btn)

        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(self.cancel_btn)
        layout.addLayout(btn_row)

        # Status / log (hidden until import starts)
        self.status_label = QLabel("")
        self.status_label.setWordWrap(True)
        layout.addWidget(self.status_label)

        self.log_text = QListWidget()
        self.log_text.setVisible(False)
        layout.addWidget(self.log_text)

        # Populate initial files
        for f in self.files:
            self.file_list.addItem(Path(f).name)

    def _add_files(self, file_paths):
        for f in file_paths:
            if f not in self.files:
                self.files.append(f)
                self.file_list.addItem(Path(f).name)

    def _browse_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "Add Files",
            "",
            "All Supported (*.xlsx *.pdf *.png *.jpg *.jpeg *.tiff *.heic *.csv *.txt);;"
            "All Files (*)",
        )
        if files:
            self._add_files(files)

    def _start_import(self):
        if not self.files:
            QMessageBox.information(self, "Import", "Add some files first.")
            return
        self._processing = True
        self.drop_zone.setVisible(False)
        self.import_btn.setEnabled(False)
        self.log_text.setVisible(True)
        from PySide6.QtCore import QTimer
        QTimer.singleShot(100, self._run_import)

    def _log(self, msg):
        # If message starts with "Generating..." or "Done —", update last line
        if msg.startswith("Generating...") or msg.startswith("Done —"):
            count = self.log_text.count()
            if count > 0:
                last = self.log_text.item(count - 1)
                if last.text().startswith("Generating...") or last.text().startswith("Sending to"):
                    last.setText(msg)
                    self.log_text.scrollToBottom()
                    QApplication.processEvents()
                    return
        self.log_text.addItem(msg)
        self.log_text.scrollToBottom()
        QApplication.processEvents()

    def _run_import(self):
        snippet_name = self.snippet_combo.currentText()
        if not snippet_name:
            self._log("No snippet selected.")
            return

        # Step 1: Extract text from all files
        self.status_label.setText("Step 1/2: Extracting text from files...")
        QApplication.processEvents()

        all_text = []
        for f in self.files:
            self._log(f"Reading: {Path(f).name}")
            text = _extract_text_from_file(f)
            is_error = (not text or text.startswith("[XLSX Error")
                       or text.startswith("[OCR Error") or text.startswith("[PDF Error")
                       or text.startswith("[Unsupported") or text.startswith("[No text"))
            if text and not is_error:
                all_text.append(text)
                self._log(f"  Extracted {len(text)} characters")
            else:
                self._log(f"  {text}")

        if not all_text:
            self._log("No text could be extracted from the files.")
            self.status_label.setText("Import failed — no text extracted.")
            return

        combined = "\n\n---\n\n".join(all_text)
        self._log(f"Total text: {len(combined)} characters")

        # Step 2: Fill snippet with AI
        self.status_label.setText(f"Step 2/2: Filling '{snippet_name}' with {OLLAMA_MODEL}...")
        QApplication.processEvents()

        result = _import_to_snippet(snippet_name, combined, log_fn=self._log)

        if result is None:
            self.status_label.setText("Import failed.")
            return

        wb, cell_data = result

        # Step 3: Prompt for missing fields one at a time
        final_data = _prompt_missing_fields(snippet_name, cell_data, self)
        _fill_snippet_cells(wb, final_data)

        self.result_wb = wb
        filled = sum(1 for v in final_data.values() if v)
        total = len(final_data)
        self._log(f"Filled {filled}/{total} fields")
        self.status_label.setText("Import complete!")
        self.cancel_btn.setText("Done")
        self.cancel_btn.clicked.disconnect()
        self.cancel_btn.clicked.connect(self.accept)


# ── Snippet Composer Dialog ─────────────────────────────────────────────────

class SnippetComposer(QDialog):
    """Pick and order snippets to build a document from blocks."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Build from Snippets")
        self.setMinimumSize(600, 450)
        self.result_wb = None

        layout = QHBoxLayout(self)

        # Left: available snippets
        left = QVBoxLayout()
        left.addWidget(QLabel("Available Snippets"))
        self.avail_list = QListWidget()
        for name in list_snippets():
            self.avail_list.addItem(name)
        left.addWidget(self.avail_list)

        add_btn = QPushButton("Add >>")
        add_btn.clicked.connect(self._add_snippet)
        left.addWidget(add_btn)
        layout.addLayout(left)

        # Right: chosen snippets in order
        right = QVBoxLayout()
        right.addWidget(QLabel("Document Order"))
        self.order_list = QListWidget()
        right.addWidget(self.order_list)

        btn_row = QHBoxLayout()
        up_btn = QPushButton("Move Up")
        up_btn.clicked.connect(self._move_up)
        btn_row.addWidget(up_btn)
        down_btn = QPushButton("Move Down")
        down_btn.clicked.connect(self._move_down)
        btn_row.addWidget(down_btn)
        remove_btn = QPushButton("Remove")
        remove_btn.clicked.connect(self._remove)
        btn_row.addWidget(remove_btn)
        right.addLayout(btn_row)

        # Bottom buttons
        bottom = QHBoxLayout()
        bottom.addStretch()
        build_btn = QPushButton("Build Document")
        build_btn.setStyleSheet(
            "QPushButton { background-color: #1A73E8; color: white; "
            "font-weight: bold; padding: 6px 20px; border-radius: 4px; }"
            "QPushButton:hover { background-color: #1557B0; }"
        )
        build_btn.clicked.connect(self._build)
        bottom.addWidget(build_btn)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        bottom.addWidget(cancel_btn)
        right.addLayout(bottom)
        layout.addLayout(right)

    def _add_snippet(self):
        item = self.avail_list.currentItem()
        if item:
            self.order_list.addItem(item.text())

    def _move_up(self):
        row = self.order_list.currentRow()
        if row > 0:
            item = self.order_list.takeItem(row)
            self.order_list.insertItem(row - 1, item)
            self.order_list.setCurrentRow(row - 1)

    def _move_down(self):
        row = self.order_list.currentRow()
        if row < self.order_list.count() - 1:
            item = self.order_list.takeItem(row)
            self.order_list.insertItem(row + 1, item)
            self.order_list.setCurrentRow(row + 1)

    def _remove(self):
        row = self.order_list.currentRow()
        if row >= 0:
            self.order_list.takeItem(row)

    def _build(self):
        if self.order_list.count() == 0:
            QMessageBox.information(self, "Build", "Add some snippets first.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet"

        current_row = 0
        for i in range(self.order_list.count()):
            name = self.order_list.item(i).text()
            rows_written = _build_snippet_to_ws(ws, name, current_row)
            current_row += rows_written

        self.result_wb = wb
        self.accept()


# ── Main Window ──────────────────────────────────────────────────────────────

class SheetEditWindow(QMainWindow):
    def __init__(self, filepath=None):
        super().__init__()
        self.setWindowTitle("SheetEdit")
        self.resize(1400, 800)
        self.setAcceptDrops(True)
        self.wb = None
        self.filepath = None
        self.rules = []
        self._dirty = False

        # Formula bar
        self._cell_label = QLabel("A1")
        self._cell_label.setObjectName("formula_label")
        self._cell_label.setFixedWidth(60)
        self._cell_label.setAlignment(Qt.AlignCenter)
        self._formula_edit = QLineEdit()
        self._formula_edit.setObjectName("formula_edit")
        self._formula_edit.setPlaceholderText("Cell contents")
        self._formula_edit.returnPressed.connect(self._formula_bar_commit)

        formula_bar = QHBoxLayout()
        formula_bar.setContentsMargins(6, 2, 6, 2)
        formula_bar.addWidget(self._cell_label)
        formula_bar.addWidget(self._formula_edit)

        self.tabs = QTabWidget()
        self.tabs.currentChanged.connect(self._tab_changed)

        central = QWidget()
        vbox = QVBoxLayout(central)
        vbox.setContentsMargins(0, 0, 0, 0)
        vbox.setSpacing(0)
        vbox.addLayout(formula_bar)
        vbox.addWidget(self.tabs)
        self.setCentralWidget(central)

        self.setStatusBar(QStatusBar())
        self.statusBar().showMessage("Ready")

        self._build_menu()
        self._build_toolbar()

        if filepath:
            self.open_file(filepath)
        else:
            self._new_workbook()

    # ── Menu bar ─────────────────────────────────────────────────────────

    def _build_menu(self):
        mb = self.menuBar()

        file_menu = mb.addMenu("File")

        new_act = QAction("New", self)
        new_act.setShortcut(QKeySequence.New)
        new_act.triggered.connect(self._new_workbook)
        file_menu.addAction(new_act)

        template_act = QAction("New from Template...", self)
        template_act.triggered.connect(self._new_from_template)
        file_menu.addAction(template_act)

        file_menu.addSeparator()

        open_act = QAction("Open...", self)
        open_act.setShortcut(QKeySequence.Open)
        open_act.triggered.connect(self._open_dialog)
        file_menu.addAction(open_act)

        save_act = QAction("Save", self)
        save_act.setShortcut(QKeySequence.Save)
        save_act.triggered.connect(self._save)
        file_menu.addAction(save_act)

        saveas_act = QAction("Save As...", self)
        saveas_act.setShortcut(QKeySequence("Ctrl+Shift+S"))
        saveas_act.triggered.connect(self._save_as)
        file_menu.addAction(saveas_act)

        save_template_act = QAction("Save as Template...", self)
        save_template_act.triggered.connect(self._save_as_template)
        file_menu.addAction(save_template_act)

        file_menu.addSeparator()

        import_act = QAction("Import Call Sheet Data...", self)
        import_act.triggered.connect(self._import_call_sheet)
        file_menu.addAction(import_act)

        compose_act = QAction("Build from Snippets...", self)
        compose_act.triggered.connect(self._snippet_compose)
        file_menu.addAction(compose_act)

        guide_act = QAction("Edit Import Guide...", self)
        guide_act.triggered.connect(self._edit_import_guide)
        file_menu.addAction(guide_act)

        self._rules_menu = QMenu("Edit Snippet Rules", self)
        self._rules_menu.aboutToShow.connect(self._populate_snippet_rules_menu)
        file_menu.addMenu(self._rules_menu)

        file_menu.addSeparator()

        preview_act = QAction("Print...", self)
        preview_act.setShortcut(QKeySequence.Print)
        preview_act.triggered.connect(self._print_preview)
        file_menu.addAction(preview_act)

        edit_menu = mb.addMenu("Edit")

        undo_act = QAction("Undo", self)
        undo_act.setShortcut(QKeySequence.Undo)
        undo_act.triggered.connect(self.cmd_undo)
        edit_menu.addAction(undo_act)

        redo_act = QAction("Redo", self)
        redo_act.setShortcut(QKeySequence.Redo)
        redo_act.triggered.connect(self.cmd_redo)
        edit_menu.addAction(redo_act)

        edit_menu.addSeparator()

        cut_act = QAction("Cut", self)
        cut_act.setShortcut(QKeySequence.Cut)
        cut_act.triggered.connect(self.cmd_cut)
        edit_menu.addAction(cut_act)

        copy_act = QAction("Copy", self)
        copy_act.setShortcut(QKeySequence.Copy)
        copy_act.triggered.connect(self.cmd_copy)
        edit_menu.addAction(copy_act)

        paste_act = QAction("Paste", self)
        paste_act.setShortcut(QKeySequence.Paste)
        paste_act.triggered.connect(self.cmd_paste)
        edit_menu.addAction(paste_act)

        edit_menu.addSeparator()

        ins_row = QAction("Insert Row", self)
        ins_row.triggered.connect(self.cmd_insert_row)
        edit_menu.addAction(ins_row)

        del_row = QAction("Delete Row", self)
        del_row.triggered.connect(self.cmd_delete_row)
        edit_menu.addAction(del_row)

        ins_col = QAction("Insert Column", self)
        ins_col.triggered.connect(self.cmd_insert_col)
        edit_menu.addAction(ins_col)

        del_col = QAction("Delete Column", self)
        del_col.triggered.connect(self.cmd_delete_col)
        edit_menu.addAction(del_col)

        edit_menu.addSeparator()

        save_snippet_act = QAction("Save Selection as Snippet...", self)
        save_snippet_act.triggered.connect(self._menu_save_snippet)
        edit_menu.addAction(save_snippet_act)

        self._insert_snippet_menu = QMenu("Insert Snippet", self)
        self._insert_snippet_menu.aboutToShow.connect(self._populate_snippet_menu)
        edit_menu.addMenu(self._insert_snippet_menu)

        # ── Rules menu ──
        rules_menu = mb.addMenu("Rules")

        edit_rules_act = QAction("Edit Rules...", self)
        edit_rules_act.triggered.connect(self._edit_rules)
        rules_menu.addAction(edit_rules_act)

        check_rules_act = QAction("Check All Rules Now", self)
        check_rules_act.triggered.connect(self._check_all_rules)
        rules_menu.addAction(check_rules_act)

        clear_rules_act = QAction("Clear All Rules", self)
        clear_rules_act.triggered.connect(self._clear_all_rules)
        rules_menu.addAction(clear_rules_act)

    # ── Toolbar ──────────────────────────────────────────────────────────

    def _build_toolbar(self):
        tb = self.addToolBar("Format")
        tb.setIconSize(QSize(20, 20))
        tb.setMovable(False)

        # Undo / Redo
        undo_btn = QAction("\u21A9", self)
        undo_btn.setToolTip("Undo (Ctrl+Z)")
        undo_btn.triggered.connect(self.cmd_undo)
        tb.addAction(undo_btn)

        redo_btn = QAction("\u21AA", self)
        redo_btn.setToolTip("Redo (Ctrl+Shift+Z)")
        redo_btn.triggered.connect(self.cmd_redo)
        tb.addAction(redo_btn)

        tb.addSeparator()

        # Bold
        bold_act = QAction("B", self)
        bold_act.setToolTip("Bold (Ctrl+B)")
        bold_act.setShortcut(QKeySequence.Bold)
        bold_act.triggered.connect(self.cmd_bold)
        f = bold_act.font()
        f.setBold(True)
        bold_act.setFont(f)
        tb.addAction(bold_act)

        # Italic
        italic_act = QAction("I", self)
        italic_act.setToolTip("Italic (Ctrl+I)")
        italic_act.setShortcut(QKeySequence.Italic)
        italic_act.triggered.connect(self.cmd_italic)
        f = italic_act.font()
        f.setItalic(True)
        italic_act.setFont(f)
        tb.addAction(italic_act)

        tb.addSeparator()

        # Fill color
        fill_act = QAction("Fill", self)
        fill_act.setToolTip("Fill Color")
        fill_act.triggered.connect(self.cmd_fill_color)
        tb.addAction(fill_act)

        # Font color
        font_act = QAction("Color", self)
        font_act.setToolTip("Font Color")
        font_act.triggered.connect(self.cmd_font_color)
        tb.addAction(font_act)

        tb.addSeparator()

        # Horizontal alignment
        for label, align in [
            ("\u2190", "left"),
            ("\u2194", "center"),
            ("\u2192", "right"),
        ]:
            act = QAction(label, self)
            act.setToolTip(f"Align {align.title()}")
            act.triggered.connect(lambda checked, a=align: self.cmd_halign(a))
            tb.addAction(act)

        tb.addSeparator()

        # Vertical alignment
        for label, align in [
            ("\u2191", "top"),
            ("\u2195", "center"),
            ("\u2193", "bottom"),
        ]:
            act = QAction(label, self)
            act.setToolTip(f"Vertical {align.title()}")
            act.triggered.connect(lambda checked, a=align: self.cmd_valign(a))
            tb.addAction(act)

        tb.addSeparator()

        # Wrap
        wrap_act = QAction("Wrap", self)
        wrap_act.setToolTip("Toggle Wrap Text")
        wrap_act.triggered.connect(self.cmd_wrap)
        tb.addAction(wrap_act)

        # Merge / Unmerge
        merge_act = QAction("Merge", self)
        merge_act.setToolTip("Merge Cells")
        merge_act.triggered.connect(self.cmd_merge)
        tb.addAction(merge_act)

        unmerge_act = QAction("Unmerge", self)
        unmerge_act.setToolTip("Unmerge Cells")
        unmerge_act.triggered.connect(self.cmd_unmerge)
        tb.addAction(unmerge_act)

        tb.addSeparator()

        # Borders
        border_act = QAction("Borders", self)
        border_act.setToolTip("Add Thin Borders to Selection")
        border_act.triggered.connect(self.cmd_borders)
        tb.addAction(border_act)

        thick_border_act = QAction("Thick", self)
        thick_border_act.setToolTip("Add Thick Borders to Selection")
        thick_border_act.triggered.connect(self.cmd_thick_borders)
        tb.addAction(thick_border_act)

        clear_border_act = QAction("No Border", self)
        clear_border_act.setToolTip("Clear Borders")
        clear_border_act.triggered.connect(self.cmd_clear_borders)
        tb.addAction(clear_border_act)

        tb.addSeparator()

        # Force text
        text_act = QAction("Force Text", self)
        text_act.setToolTip("Force selected cells to text format")
        text_act.triggered.connect(self.cmd_force_text)
        tb.addAction(text_act)

        tb.addSeparator()

        # Insert Snippet
        insert_snip_act = QAction("Insert Snippet", self)
        insert_snip_act.setToolTip("Insert a snippet at the current cell")
        insert_snip_act.triggered.connect(self._toolbar_insert_snippet)
        tb.addAction(insert_snip_act)

        # Snippet Rules
        rules_act = QAction("Snippet Rules", self)
        rules_act.setToolTip("Edit import rules for snippets")
        rules_act.triggered.connect(self._open_snippet_rules_picker)
        tb.addAction(rules_act)

    # ── File I/O ─────────────────────────────────────────────────────────

    def _check_unsaved(self):
        """If there are unsaved changes, prompt the user. Returns True if safe to proceed."""
        if not self._dirty:
            return True
        reply = QMessageBox.question(
            self, "Unsaved Changes",
            "You have unsaved changes. Save before continuing?",
            QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
        )
        if reply == QMessageBox.Save:
            self._save()
            return True
        elif reply == QMessageBox.Discard:
            return True
        return False

    def _new_workbook(self):
        if not self._check_unsaved():
            return
        self.wb = openpyxl.Workbook()
        self.filepath = None
        self.rules = []
        self._dirty = False
        self.setWindowTitle("SheetEdit — New Workbook")
        self._reload_tabs()

    def _new_from_template(self):
        if not self._check_unsaved():
            return
        dlg = TemplatePicker(self)
        if dlg.exec() == QDialog.Accepted and dlg.chosen_wb:
            self.wb = dlg.chosen_wb
            self.filepath = None
            self.rules = []
            # Check for companion rules in templates dir
            if hasattr(dlg, '_chosen_path') and dlg._chosen_path:
                rp = Path(dlg._chosen_path).with_suffix('.rules.json')
                if rp.exists():
                    try:
                        self.rules = json.loads(rp.read_text())
                    except Exception:
                        pass
            self.setWindowTitle("SheetEdit — New from Template")
            self._reload_tabs()
            self._apply_rule_markers()
            self.statusBar().showMessage("Created from template")

    def _save_as_template(self):
        if not self.wb:
            return
        _ensure_templates_dir()
        name, ok = QInputDialog.getText(self, "Save as Template", "Template name:")
        if not ok or not name.strip():
            return
        name = name.strip()
        dest = TEMPLATES_DIR / f"{name}.xlsx"
        if dest.exists():
            reply = QMessageBox.question(
                self, "Overwrite Template",
                f"Template \"{name}\" already exists. Overwrite?",
                QMessageBox.Yes | QMessageBox.No,
            )
            if reply != QMessageBox.Yes:
                return
        try:
            for i in range(self.tabs.count()):
                self.tabs.widget(i).sync_to_ws()
            self.wb.save(str(dest))
            # Save companion rules if any
            rules_dest = dest.with_suffix('.rules.json')
            if self.rules:
                rules_dest.write_text(json.dumps(self.rules, indent=2))
            elif rules_dest.exists():
                rules_dest.unlink()
            self.statusBar().showMessage(f"Saved template: {name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save template:\n{e}")

    def _import_call_sheet(self, files=None):
        dlg = ImportCallSheetDialog(files, self)
        if dlg.exec() == QDialog.Accepted and dlg.result_wb:
            self.wb = dlg.result_wb
            self.filepath = None
            self._dirty = True
            self.setWindowTitle("SheetEdit — Imported Call Sheet")
            self._reload_tabs()
            self.statusBar().showMessage(
                f"Imported call sheet from {len(dlg.files)} file(s)"
            )

    def _snippet_compose(self):
        dlg = SnippetComposer(self)
        if dlg.exec() == QDialog.Accepted and dlg.result_wb:
            self.wb = dlg.result_wb
            self.filepath = None
            self._dirty = True
            self.setWindowTitle("SheetEdit — Composed Document")
            self._reload_tabs()
            self.statusBar().showMessage("Built document from snippets")

    def _populate_snippet_rules_menu(self):
        self._rules_menu.clear()
        snippets = list_snippets()
        if not snippets:
            act = QAction("(no snippets)", self)
            act.setEnabled(False)
            self._rules_menu.addAction(act)
            return
        for name in snippets:
            act = QAction(name, self)
            has_rules = _snippet_guide_path(name).exists()
            if has_rules:
                act.setText(f"{name}  ✓")
            act.triggered.connect(lambda checked, n=name: self._open_snippet_rules(n))
            self._rules_menu.addAction(act)
        self._rules_menu.addSeparator()
        compile_act = QAction("Compile All Rules Now", self)
        compile_act.triggered.connect(self._compile_all_rules)
        self._rules_menu.addAction(compile_act)

    def _toolbar_insert_snippet(self):
        sv = self._sheet()
        if sv is None:
            return
        snippets = list_snippets()
        if not snippets:
            QMessageBox.information(self, "Insert Snippet", "No snippets saved yet.")
            return
        name, ok = QInputDialog.getItem(
            self, "Insert Snippet", "Choose a snippet:", snippets, 0, False
        )
        if ok and name:
            row, col = sv.currentRow(), sv.currentColumn()
            insert_snippet(name, sv, row, col)
            sv.scrollToItem(sv.item(row, col))
            self.statusBar().showMessage(f"Inserted snippet: {name}")

    def _open_snippet_rules_picker(self):
        self._open_snippet_rules()

    def _open_snippet_rules(self, snippet_name=None):
        snippets = list_snippets()
        if not snippets:
            QMessageBox.information(self, "Snippet Rules", "No snippets saved yet.")
            return
        dlg = SnippetRulesEditor(snippet_name, self)
        if dlg.exec() == QDialog.Accepted:
            self.statusBar().showMessage("Import rules saved — guide compiled")

    def _compile_all_rules(self):
        _compile_import_guide()
        self.statusBar().showMessage("Import guide compiled from all snippet rules")

    def _edit_import_guide(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Import Guide")
        dlg.resize(700, 600)
        layout = QVBoxLayout(dlg)

        info = QLabel("This guide tells the AI how to map imported data to your call sheet snippets.")
        info.setWordWrap(True)
        layout.addWidget(info)

        editor = QPlainTextEdit()
        editor.setFont(QFont("Menlo", 12))
        editor.setLineWrapMode(QPlainTextEdit.NoWrap)
        # Load existing guide
        if IMPORT_GUIDE_PATH.exists():
            editor.setPlainText(IMPORT_GUIDE_PATH.read_text())
        else:
            editor.setPlainText("# Import Guide\n\nNo guide file found. Add instructions here.\n")
        layout.addWidget(editor)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        save_btn = QPushButton("Save")
        save_btn.setStyleSheet(
            "QPushButton { background-color: #1A73E8; color: white; "
            "font-weight: bold; padding: 6px 20px; border-radius: 4px; }"
            "QPushButton:hover { background-color: #1557B0; }"
        )
        save_btn.clicked.connect(lambda: self._save_import_guide(editor.toPlainText(), dlg))
        btn_row.addWidget(save_btn)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(dlg.reject)
        btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)

        dlg.exec()

    def _save_import_guide(self, text, dlg):
        IMPORT_GUIDE_PATH.parent.mkdir(parents=True, exist_ok=True)
        IMPORT_GUIDE_PATH.write_text(text)
        self.statusBar().showMessage("Import guide saved")
        dlg.accept()

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                if Path(url.toLocalFile()).suffix.lower() in SUPPORTED_IMPORT_EXTS:
                    event.acceptProposedAction()
                    return

    def dropEvent(self, event):
        files = []
        for url in event.mimeData().urls():
            path = url.toLocalFile()
            if path and Path(path).suffix.lower() in SUPPORTED_IMPORT_EXTS:
                files.append(path)
        if not files:
            return
        # A single .xlsx always opens as a spreadsheet — never auto-import.
        if len(files) == 1 and Path(files[0]).suffix.lower() == ".xlsx":
            self.open_file(files[0])
            return
        self._import_call_sheet(files)

    def _open_dialog(self):
        if not self._check_unsaved():
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Spreadsheet", "", "Excel Files (*.xlsx);;All Files (*)"
        )
        if path:
            self.open_file(path)

    def open_file(self, path):
        try:
            self.wb = openpyxl.load_workbook(path)
            self.filepath = path
            self.rules = []
            rp = self._rules_path()
            if rp and rp.exists():
                try:
                    self.rules = json.loads(rp.read_text())
                except Exception:
                    pass
            self.setWindowTitle(f"SheetEdit \u2014 {Path(path).name}")
            self._reload_tabs()
            self._apply_rule_markers()
            self.statusBar().showMessage(f"Opened {Path(path).name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open file:\n{e}")

    def _reload_tabs(self):
        self.tabs.clear()
        for name in self.wb.sheetnames:
            sv = SheetView(self.wb[name], self)
            sv.currentCellChanged.connect(self._on_cell_changed)
            self.tabs.addTab(sv, name)

    def _tab_changed(self, index):
        self._update_formula_bar()

    def _on_cell_changed(self, row, col, prev_row, prev_col):
        self._update_formula_bar()

    def _update_formula_bar(self):
        sv = self._sheet()
        if sv is None:
            self._cell_label.setText("")
            self._formula_edit.setText("")
            return
        row = sv.currentRow()
        col = sv.currentColumn()
        if row < 0 or col < 0:
            self._cell_label.setText("")
            self._formula_edit.setText("")
            return
        ref = f"{get_column_letter(col + 1)}{row + 1}"
        self._cell_label.setText(ref)
        item = sv.item(row, col)
        if item:
            self._formula_edit.setText(item.text())
        else:
            self._formula_edit.setText("")

    def _formula_bar_commit(self):
        sv = self._sheet()
        if sv is None:
            return
        row = sv.currentRow()
        col = sv.currentColumn()
        item = sv.item(row, col)
        if item is None:
            item = QTableWidgetItem()
            sv.setItem(row, col, item)
        sv.push_undo([(row, col)])
        item.setText(self._formula_edit.text())

    def _save(self):
        if self.filepath:
            self._write(self.filepath)
        else:
            self._save_as()

    def _save_as(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Spreadsheet", "", "Excel Files (*.xlsx)"
        )
        if path:
            if not path.endswith(".xlsx"):
                path += ".xlsx"
            self.filepath = path
            self._write(path)

    def _write(self, path):
        try:
            for i in range(self.tabs.count()):
                sv = self.tabs.widget(i)
                sv.sync_to_ws()
            self.wb.save(path)
            # Save/remove rules sidecar
            rp = self._rules_path()
            if rp:
                if self.rules:
                    rp.write_text(json.dumps(self.rules, indent=2))
                elif rp.exists():
                    rp.unlink()
            self._dirty = False
            self.setWindowTitle(f"SheetEdit \u2014 {Path(path).name}")
            self.statusBar().showMessage(f"Saved to {Path(path).name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save:\n{e}")

    def closeEvent(self, event):
        if self._check_unsaved():
            event.accept()
        else:
            event.ignore()

    # ── Rules helpers ────────────────────────────────────────────────────

    def _rules_path(self):
        """Return the .rules.json sidecar path for the current file."""
        if not self.filepath:
            return None
        return Path(self.filepath).with_suffix('.rules.json')

    def _apply_rule_markers(self):
        """Mark cells that are in ruled ranges with RULE_ROLE for the blue dot."""
        sv = self._sheet()
        if sv is None:
            return
        # Clear all rule markers first
        for ri in range(sv.rowCount()):
            for ci in range(sv.columnCount()):
                item = sv.item(ri, ci)
                if item:
                    item.setData(RULE_ROLE, None)
        # Set markers for ruled cells
        for entry in self.rules:
            try:
                min_c, min_r, max_c, max_r = range_boundaries(entry["range"])
            except Exception:
                continue
            for ri in range(min_r - 1, min(max_r, sv.rowCount())):
                for ci in range(min_c - 1, min(max_c, sv.columnCount())):
                    item = sv.item(ri, ci)
                    if item is None:
                        item = QTableWidgetItem()
                        sv.setItem(ri, ci, item)
                    item.setData(RULE_ROLE, True)
        sv.viewport().update()

    def _check_rules(self, cells):
        """Check rules for a list of (row, col) cells. Returns list of violation strings."""
        sv = self._sheet()
        if sv is None or not self.rules:
            return []
        violations = []
        cell_set = set(cells)
        for entry in self.rules:
            try:
                min_c, min_r, max_c, max_r = range_boundaries(entry["range"])
            except Exception:
                continue
            for rule in entry.get("rules", []):
                rtype = rule.get("type", "")
                for ri in range(min_r - 1, max_r):
                    for ci in range(min_c - 1, max_c):
                        if (ri, ci) not in cell_set:
                            continue
                        item = sv.item(ri, ci)
                        text = item.text() if item else ""
                        cell_ref = f"{get_column_letter(ci + 1)}{ri + 1}"

                        if rtype == "not_empty" and not text.strip():
                            violations.append(f"{cell_ref}: must not be empty")
                        elif rtype == "number_only" and text.strip():
                            try:
                                float(text)
                            except ValueError:
                                violations.append(f"{cell_ref}: must be a number")
                        elif rtype == "max_length":
                            ml = rule.get("max_length", 999999)
                            if len(text) > ml:
                                violations.append(f"{cell_ref}: text exceeds {ml} chars")
                        elif rtype == "font_size_min" and item:
                            sz = item.font().pointSize()
                            mn = rule.get("min_size", 0)
                            if sz < mn:
                                violations.append(f"{cell_ref}: font size {sz} < minimum {mn}")
                        elif rtype == "font_size_max" and item:
                            sz = item.font().pointSize()
                            mx = rule.get("max_size", 999)
                            if sz > mx:
                                violations.append(f"{cell_ref}: font size {sz} > maximum {mx}")
                        elif rtype == "wrap_required" and item:
                            if not (item.textAlignment() & Qt.TextWordWrap):
                                violations.append(f"{cell_ref}: wrap text required")
                        elif rtype == "fill_required" and item:
                            bg = item.background()
                            if bg == QBrush() or not bg.color().isValid() or bg.color() == QColor("#FFFFFF"):
                                violations.append(f"{cell_ref}: background fill required")
                        elif rtype == "row_height":
                            expected = rule.get("height", 0)
                            actual_px = sv.rowHeight(ri)
                            actual_pt = px_to_xl_row_height(actual_px)
                            if abs(actual_pt - expected) > 1.0:
                                violations.append(f"Row {ri + 1}: height {actual_pt:.0f}pt != expected {expected}pt")
        return violations

    def _check_and_warn(self, cells):
        """Check rules for cells and warn user if violations found."""
        violations = self._check_rules(cells)
        if violations:
            msg = "\n".join(violations[:20])
            if len(violations) > 20:
                msg += f"\n... and {len(violations) - 20} more"
            QMessageBox.warning(self, "Rule Violations", msg)

    def _check_and_warn_selection(self):
        """Check rules for the currently selected cells."""
        coords = self._selected_coords()
        if coords:
            self._check_and_warn(coords)

    # ── Snippet menu actions ─────────────────────────────────────────────

    def _menu_save_snippet(self):
        sv = self._sheet()
        if sv is None:
            return
        sv._ctx_save_snippet()

    def _populate_snippet_menu(self):
        self._insert_snippet_menu.clear()
        snippets = list_snippets()
        if not snippets:
            act = QAction("(no snippets)", self)
            act.setEnabled(False)
            self._insert_snippet_menu.addAction(act)
            return
        sv = self._sheet()
        for name in snippets:
            act = QAction(name, self)
            act.triggered.connect(lambda checked, n=name: self._menu_insert_snippet(n))
            self._insert_snippet_menu.addAction(act)

    def _menu_insert_snippet(self, name):
        sv = self._sheet()
        if sv is None:
            return
        sv._ctx_insert_snippet(name)

    # ── Rules menu actions ───────────────────────────────────────────────

    def _edit_rules(self):
        dlg = RulesEditor(self.rules, self)
        if dlg.exec() == QDialog.Accepted:
            self.rules = dlg.get_rules()
            self._apply_rule_markers()

    def _check_all_rules(self):
        sv = self._sheet()
        if sv is None:
            return
        all_cells = [(r, c) for r in range(sv.rowCount()) for c in range(sv.columnCount())]
        violations = self._check_rules(all_cells)
        if violations:
            msg = "\n".join(violations[:50])
            if len(violations) > 50:
                msg += f"\n... and {len(violations) - 50} more"
            QMessageBox.warning(self, "Rule Violations", f"{len(violations)} violation(s) found:\n\n{msg}")
        else:
            QMessageBox.information(self, "Rules", "All rules pass.")

    def _clear_all_rules(self):
        if not self.rules:
            QMessageBox.information(self, "Rules", "No rules to clear.")
            return
        reply = QMessageBox.question(
            self, "Clear Rules",
            f"Remove all {len(self.rules)} rule entries?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            self.rules = []
            self._apply_rule_markers()
            self.statusBar().showMessage("Rules cleared")

    # ── Print ────────────────────────────────────────────────────────────

    def _auto_orientation(self, sv):
        """Default to portrait orientation."""
        return QPageLayout.Portrait

    def _print_preview(self):
        from PySide6.QtPrintSupport import QPrintPreviewWidget
        sv = self._sheet()
        if sv is None:
            return
        printer = QPrinter(QPrinter.ScreenResolution)
        printer.setPageOrientation(self._auto_orientation(sv))
        dialog = QPrintPreviewDialog(printer, self)
        dialog.paintRequested.connect(lambda p: self._render_sheet(p, sv))

        # Find the preview widget inside the dialog
        preview = dialog.findChild(QPrintPreviewWidget)

        # Customize the toolbar
        toolbars = dialog.findChildren(QToolBar)
        if toolbars and preview:
            tb = toolbars[0]

            # Remove the default print button (first action)
            actions = tb.actions()
            if actions:
                tb.removeAction(actions[0])

            tb.addSeparator()
            portrait_act = QAction("Portrait", dialog)
            portrait_act.triggered.connect(preview.setPortraitOrientation)
            tb.addAction(portrait_act)
            landscape_act = QAction("Landscape", dialog)
            landscape_act.triggered.connect(preview.setLandscapeOrientation)
            tb.addAction(landscape_act)

            # Add spacer to push PRINT to the right
            spacer = QWidget()
            spacer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
            tb.addWidget(spacer)

            print_btn = QPushButton("PRINT")
            print_btn.setStyleSheet(
                "QPushButton { background-color: #1A73E8; color: white; "
                "font-weight: bold; padding: 6px 20px; border-radius: 4px; "
                "font-size: 13px; }"
                "QPushButton:hover { background-color: #1557B0; }"
            )
            print_btn.clicked.connect(lambda: (dialog.accept(), self._do_print(printer, sv)))
            tb.addWidget(print_btn)

        dialog.exec()

    def _do_print(self, printer, sv):
        """Print using the already-configured printer from preview."""
        dialog = QPrintDialog(printer, self)
        if dialog.exec() == QPrintDialog.Accepted:
            self._render_sheet(printer, sv)

    def _render_sheet(self, printer, sv: "SheetView"):
        """Render the sheet: fit width to page, paginate rows across pages."""
        painter = QPainter()
        if not painter.begin(printer):
            return

        # Use painter coordinate space (matches what drawRect/drawText use)
        pw = float(painter.device().width())
        ph = float(painter.device().height())

        # Find used range
        max_row = sv.rowCount()
        max_col = sv.columnCount()

        def _cell_has_content(item):
            if item is None:
                return False
            if item.text():
                return True
            bg = item.background()
            if bg != QBrush() and bg.color().isValid() and bg.color() != QColor("#FFFFFF"):
                return True
            if item.data(BORDER_ROLE):
                return True
            return False

        # Trim trailing empty rows
        for ri in range(max_row - 1, -1, -1):
            for ci in range(max_col):
                if _cell_has_content(sv.item(ri, ci)):
                    max_row = ri + 1
                    break
            else:
                continue
            break
        else:
            max_row = 1

        # Trim trailing empty columns
        for ci in range(max_col - 1, -1, -1):
            for ri in range(max_row):
                if _cell_has_content(sv.item(ri, ci)):
                    max_col = ci + 1
                    break
            else:
                continue
            break
        else:
            max_col = 1

        col_widths = [sv.columnWidth(c) for c in range(max_col)]
        row_heights = [sv.rowHeight(r) for r in range(max_row)]

        total_w = sum(col_widths)
        if total_w <= 0:
            painter.end()
            return

        # Fit WIDTH to page — scale everything proportionally
        scale = pw / total_w

        # Paginate: figure out which rows fit on each page
        pages = []
        page_start = 0
        page_y = 0.0
        page_h = ph / scale  # page height in screen-pixel units

        for ri in range(max_row):
            rh = row_heights[ri]
            if page_y + rh > page_h and ri > page_start:
                pages.append((page_start, ri - 1))
                page_start = ri
                page_y = rh
            else:
                page_y += rh
        pages.append((page_start, max_row - 1))

        # Build merge lookup: (row, col) -> (r1, c1, r2, c2) for the top-left cell
        # Also track which cells are "hidden" (inside a merge but not the origin)
        merge_origins = {}  # (r1, c1) -> (r1, c1, r2, c2)
        merged_hidden = set()  # cells that are covered by a merge
        for m in sv.merges:
            r1, c1, r2, c2 = m
            merge_origins[(r1, c1)] = m
            for mr in range(r1, r2 + 1):
                for mc in range(c1, c2 + 1):
                    if (mr, mc) != (r1, c1):
                        merged_hidden.add((mr, mc))

        # Precompute cumulative column x positions
        col_x = [0.0]
        for cw in col_widths:
            col_x.append(col_x[-1] + cw)

        for page_idx, (row_start, row_end) in enumerate(pages):
            if page_idx > 0:
                printer.newPage()

            painter.save()
            painter.scale(scale, scale)

            # Precompute row y positions relative to page
            row_y = [0.0]
            for ri in range(row_start, row_end + 1):
                row_y.append(row_y[-1] + row_heights[ri])

            # First pass: draw grid lines and fills for all cells
            for ri_idx, ri in enumerate(range(row_start, row_end + 1)):
                for ci in range(max_col):
                    if (ri, ci) in merged_hidden:
                        continue

                    x = col_x[ci]
                    y = row_y[ri_idx]

                    # Check if this is a merge origin
                    if (ri, ci) in merge_origins:
                        mr1, mc1, mr2, mc2 = merge_origins[(ri, ci)]
                        cw = col_x[min(mc2 + 1, max_col)] - col_x[mc1]
                        # Row height for merge: sum heights of spanned rows on this page
                        rh = 0.0
                        for mr in range(mr1, mr2 + 1):
                            if row_start <= mr <= row_end:
                                rh += row_heights[mr]
                    else:
                        cw = col_widths[ci]
                        rh = row_heights[ri]

                    cell_rect = QRectF(x, y, cw, rh)

                    item = sv.item(ri, ci)
                    if item:
                        # Fill
                        bg = item.background()
                        if bg != QBrush() and bg.color().isValid() and bg.color() != QColor("#FFFFFF"):
                            painter.fillRect(cell_rect, bg)

                        # Borders
                        bd = item.data(BORDER_ROLE)
                        if bd:
                            for side_name, coords in [
                                ("top", (x, y, x + cw, y)),
                                ("bottom", (x, y + rh, x + cw, y + rh)),
                                ("left", (x, y, x, y + rh)),
                                ("right", (x + cw, y, x + cw, y + rh)),
                            ]:
                                info = bd.get(side_name)
                                if info:
                                    color_hex, bw, style = info
                                    pen = QPen(QColor(color_hex), max(bw, 1), style)
                                    painter.setPen(pen)
                                    painter.drawLine(
                                        QPointF(coords[0], coords[1]),
                                        QPointF(coords[2], coords[3]),
                                    )

                        # Text
                        text = item.text()
                        if text:
                            qf = QFont(item.font())
                            painter.setFont(qf)

                            fg = item.foreground()
                            if fg != QBrush():
                                painter.setPen(QPen(fg.color()))
                            else:
                                painter.setPen(QPen(QColor("#202124")))

                            flags = item.textAlignment()
                            text_rect = cell_rect.adjusted(3, 1, -3, -1)
                            painter.drawText(text_rect, flags, text)

                    # Grid line
                    painter.setPen(QPen(QColor("#D0D0D0"), 0.5))
                    painter.drawRect(cell_rect)

            painter.restore()

        painter.end()

    # ── Selection helpers ────────────────────────────────────────────────

    def _sheet(self) -> SheetView:
        return self.tabs.currentWidget()

    def _selected_items(self):
        sv = self._sheet()
        if sv is None:
            return []
        return sv.selectedItems()

    def _selected_coords(self):
        """Return list of (row, col) for selected cells."""
        sv = self._sheet()
        if sv is None:
            return []
        return [(item.row(), item.column()) for item in sv.selectedItems()]

    def _push_undo_for_selection(self):
        sv = self._sheet()
        if sv is None:
            return
        coords = self._selected_coords()
        if coords:
            sv.push_undo(coords)

    def cmd_undo(self):
        sv = self._sheet()
        if sv:
            sv.undo()

    def cmd_redo(self):
        sv = self._sheet()
        if sv:
            sv.redo()

    def cmd_copy(self):
        sv = self._sheet()
        if sv:
            sv.copy_cells()

    def cmd_cut(self):
        sv = self._sheet()
        if sv:
            sv.cut_cells()

    def cmd_paste(self):
        sv = self._sheet()
        if sv:
            sv.paste_cells()

    def _selected_range(self):
        sv = self._sheet()
        if sv is None:
            return None
        ranges = sv.selectedRanges()
        if not ranges:
            return None
        r = ranges[0]
        return (r.topRow(), r.leftColumn(), r.bottomRow(), r.rightColumn())

    # ── Format commands ──────────────────────────────────────────────────

    def cmd_bold(self):
        self._push_undo_for_selection()
        for item in self._selected_items():
            f = item.font()
            f.setBold(not f.bold())
            item.setFont(f)
        self._check_and_warn_selection()

    def cmd_italic(self):
        self._push_undo_for_selection()
        for item in self._selected_items():
            f = item.font()
            f.setItalic(not f.italic())
            item.setFont(f)
        self._check_and_warn_selection()

    def cmd_fill_color(self):
        color = QColorDialog.getColor(QColor(Qt.white), self, "Fill Color")
        if color.isValid():
            self._push_undo_for_selection()
            for item in self._selected_items():
                item.setBackground(QBrush(color))
            self._check_and_warn_selection()

    def cmd_font_color(self):
        color = QColorDialog.getColor(QColor(Qt.black), self, "Font Color")
        if color.isValid():
            self._push_undo_for_selection()
            for item in self._selected_items():
                item.setForeground(QBrush(color))
            self._check_and_warn_selection()

    def cmd_halign(self, align):
        self._push_undo_for_selection()
        flag = HALIGN_MAP[align]
        for item in self._selected_items():
            cur = item.textAlignment()
            cur &= ~(Qt.AlignLeft | Qt.AlignHCenter | Qt.AlignRight)
            item.setTextAlignment(cur | flag)
        self._check_and_warn_selection()

    def cmd_valign(self, align):
        self._push_undo_for_selection()
        flag = VALIGN_MAP[align]
        for item in self._selected_items():
            cur = item.textAlignment()
            cur &= ~(Qt.AlignTop | Qt.AlignVCenter | Qt.AlignBottom)
            item.setTextAlignment(cur | flag)
        self._check_and_warn_selection()

    def cmd_wrap(self):
        self._push_undo_for_selection()
        for item in self._selected_items():
            cur = item.textAlignment()
            item.setTextAlignment(cur ^ Qt.TextWordWrap)
        self._check_and_warn_selection()

    def cmd_merge(self):
        rng = self._selected_range()
        if rng is None:
            return
        r1, c1, r2, c2 = rng
        if r1 == r2 and c1 == c2:
            return
        sv = self._sheet()
        sv.setSpan(r1, c1, r2 - r1 + 1, c2 - c1 + 1)
        sv.merges.append((r1, c1, r2, c2))
        self.statusBar().showMessage(
            f"Merged {get_column_letter(c1+1)}{r1+1}:{get_column_letter(c2+1)}{r2+1}"
        )

    def cmd_unmerge(self):
        rng = self._selected_range()
        if rng is None:
            return
        r1, c1, r2, c2 = rng
        sv = self._sheet()
        to_remove = []
        for m in sv.merges:
            if m[0] <= r2 and m[2] >= r1 and m[1] <= c2 and m[3] >= c1:
                to_remove.append(m)
        if not to_remove:
            return
        # Remove from our list
        for m in to_remove:
            sv.merges.remove(m)
        # Clear ALL spans at once (safe single Qt call), then re-apply survivors
        sv.setUpdatesEnabled(False)
        sv.clearSpans()
        for m in sv.merges:
            sv.setSpan(m[0], m[1], m[2] - m[0] + 1, m[3] - m[1] + 1)
        sv.setUpdatesEnabled(True)
        self.statusBar().showMessage("Unmerged")

    def _apply_borders(self, color_hex, width):
        """Apply border to all selected cells."""
        self._push_undo_for_selection()
        rng = self._selected_range()
        if rng is None:
            return
        r1, c1, r2, c2 = rng
        sv = self._sheet()
        style = Qt.SolidLine
        for ri in range(r1, r2 + 1):
            for ci in range(c1, c2 + 1):
                item = sv.item(ri, ci)
                if item is None:
                    item = QTableWidgetItem()
                    sv.setItem(ri, ci, item)
                bd = item.data(BORDER_ROLE) or {}
                # Top edge
                if ri == r1:
                    bd["top"] = (color_hex, width, style)
                # Bottom edge
                if ri == r2:
                    bd["bottom"] = (color_hex, width, style)
                # Left edge
                if ci == c1:
                    bd["left"] = (color_hex, width, style)
                # Right edge
                if ci == c2:
                    bd["right"] = (color_hex, width, style)
                # Inner borders
                if ri < r2:
                    bd["bottom"] = (color_hex, width, style)
                if ri > r1:
                    bd["top"] = (color_hex, width, style)
                if ci < c2:
                    bd["right"] = (color_hex, width, style)
                if ci > c1:
                    bd["left"] = (color_hex, width, style)
                item.setData(BORDER_ROLE, bd)
        sv.viewport().update()
        self.statusBar().showMessage("Borders applied")

    def cmd_borders(self):
        self._apply_borders("#BFBFBF", 1)

    def cmd_thick_borders(self):
        self._apply_borders("#000000", 2)

    def cmd_clear_borders(self):
        self._push_undo_for_selection()
        for item in self._selected_items():
            item.setData(BORDER_ROLE, None)
        sv = self._sheet()
        if sv:
            sv.viewport().update()
        self.statusBar().showMessage("Borders cleared")

    def cmd_force_text(self):
        self._push_undo_for_selection()
        for item in self._selected_items():
            txt = item.text()
            item.setText(txt)
        self.statusBar().showMessage("Cells forced to text")

    # ── Row/Column insert/delete ─────────────────────────────────────────

    def cmd_insert_row(self):
        sv = self._sheet()
        if sv is None:
            return
        row = sv.currentRow()
        # Clear existing spans before insert so Qt doesn't reference stale data
        sv.setUpdatesEnabled(False)
        for m in sv.merges:
            sv.setSpan(m[0], m[1], 1, 1)
        sv.insertRow(row)
        # Shift merges below insertion point
        sv.merges = [
            (r1 + (1 if r1 >= row else 0), c1, r2 + (1 if r2 >= row else 0), c2)
            for r1, c1, r2, c2 in sv.merges
        ]
        # Re-apply adjusted merges
        for r1, c1, r2, c2 in sv.merges:
            if r2 > r1 or c2 > c1:
                sv.setSpan(r1, c1, r2 - r1 + 1, c2 - c1 + 1)
        sv.setUpdatesEnabled(True)
        self.statusBar().showMessage(f"Inserted row at {row + 1}")

    def cmd_delete_row(self):
        sv = self._sheet()
        if sv is None:
            return
        row = sv.currentRow()
        # Clean up merges that overlap this row
        sv.setUpdatesEnabled(False)
        to_remove = [m for m in sv.merges if m[0] <= row <= m[2]]
        for m in to_remove:
            sv.setSpan(m[0], m[1], 1, 1)
            sv.merges.remove(m)
        # Shift remaining merges above deleted row
        sv.merges = [
            (r1 - (1 if r1 > row else 0), c1, r2 - (1 if r2 > row else 0), c2)
            for r1, c1, r2, c2 in sv.merges
            if not (r1 > row and r2 < row)  # shouldn't happen, but guard
        ]
        sv.removeRow(row)
        # Re-apply adjusted merges
        for r1, c1, r2, c2 in sv.merges:
            if r2 > r1 or c2 > c1:
                sv.setSpan(r1, c1, r2 - r1 + 1, c2 - c1 + 1)
        sv.setUpdatesEnabled(True)
        self.statusBar().showMessage(f"Deleted row {row + 1}")

    def cmd_insert_col(self):
        sv = self._sheet()
        if sv is None:
            return
        col = sv.currentColumn()
        # Clear existing spans before insert so Qt doesn't reference stale data
        sv.setUpdatesEnabled(False)
        for m in sv.merges:
            sv.setSpan(m[0], m[1], 1, 1)
        sv.insertColumn(col)
        # Shift merges past insertion point
        sv.merges = [
            (r1, c1 + (1 if c1 >= col else 0), r2, c2 + (1 if c2 >= col else 0))
            for r1, c1, r2, c2 in sv.merges
        ]
        # Re-apply adjusted merges
        for r1, c1, r2, c2 in sv.merges:
            if r2 > r1 or c2 > c1:
                sv.setSpan(r1, c1, r2 - r1 + 1, c2 - c1 + 1)
        sv.setUpdatesEnabled(True)
        self.statusBar().showMessage(
            f"Inserted column at {get_column_letter(col + 1)}"
        )

    def cmd_delete_col(self):
        sv = self._sheet()
        if sv is None:
            return
        col = sv.currentColumn()
        # Clean up merges that overlap this column
        sv.setUpdatesEnabled(False)
        to_remove = [m for m in sv.merges if m[1] <= col <= m[3]]
        for m in to_remove:
            sv.setSpan(m[0], m[1], 1, 1)
            sv.merges.remove(m)
        # Shift remaining merges past deleted column
        sv.merges = [
            (r1, c1 - (1 if c1 > col else 0), r2, c2 - (1 if c2 > col else 0))
            for r1, c1, r2, c2 in sv.merges
        ]
        sv.removeColumn(col)
        # Re-apply adjusted merges
        for r1, c1, r2, c2 in sv.merges:
            if r2 > r1 or c2 > c1:
                sv.setSpan(r1, c1, r2 - r1 + 1, c2 - c1 + 1)
        sv.setUpdatesEnabled(True)
        self.statusBar().showMessage(
            f"Deleted column {get_column_letter(col + 1)}"
        )


# ── Stylesheet ───────────────────────────────────────────────────────────────

LIGHT_STYLESHEET = """
QMainWindow, QWidget {
    background-color: #FFFFFF;
    color: #202124;
}
QMenuBar {
    background-color: #F8F9FA;
    border-bottom: 1px solid #DADCE0;
    color: #202124;
    padding: 2px;
}
QMenuBar::item:selected {
    background-color: #E8EAED;
    border-radius: 4px;
}
QMenu {
    background-color: #FFFFFF;
    border: 1px solid #DADCE0;
    color: #202124;
}
QMenu::item:selected {
    background-color: #E8F0FE;
}
QToolBar {
    background-color: #F1F3F4;
    border-bottom: 1px solid #DADCE0;
    spacing: 2px;
    padding: 3px 6px;
}
QToolBar QToolButton {
    background-color: transparent;
    border: 1px solid transparent;
    border-radius: 4px;
    padding: 4px 8px;
    color: #444746;
    font-size: 13px;
}
QToolBar QToolButton:hover {
    background-color: #E8EAED;
    border: 1px solid #DADCE0;
}
QToolBar QToolButton:pressed {
    background-color: #D2E3FC;
}
QToolBar::separator {
    width: 1px;
    background-color: #DADCE0;
    margin: 4px 4px;
}
QToolBarExtension {
    background-color: #F1F3F4;
    border: none;
    padding: 2px;
}
QTabWidget::pane {
    border: none;
    background-color: #FFFFFF;
}
QTabBar::tab {
    background-color: #F1F3F4;
    border: 1px solid #DADCE0;
    border-bottom: none;
    padding: 6px 16px;
    margin-right: 2px;
    border-top-left-radius: 4px;
    border-top-right-radius: 4px;
    color: #5F6368;
    font-size: 12px;
}
QTabBar::tab:selected {
    background-color: #FFFFFF;
    color: #1A73E8;
    font-weight: bold;
    border-bottom: 2px solid #1A73E8;
}
QTabBar::tab:hover:!selected {
    background-color: #E8EAED;
}
QTableWidget {
    background-color: #FFFFFF;
    gridline-color: #E2E2E2;
    border: none;
    selection-background-color: #D2E3FC;
    selection-color: #202124;
    color: #202124;
    font-family: Arial;
    font-size: 13px;
}
QTableWidget::item {
    padding: 1px 3px;
}
QTableWidget::item:selected {
    background-color: #D2E3FC;
    color: #202124;
}
QTableWidget QTableCornerButton::section {
    background-color: #F8F9FA;
    border: 1px solid #E2E2E2;
}
QHeaderView::section {
    background-color: #F8F9FA;
    color: #5F6368;
    border: 1px solid #E2E2E2;
    padding: 4px;
    font-family: Arial;
    font-size: 12px;
    font-weight: 500;
}
QHeaderView::section:checked {
    background-color: #D2E3FC;
    color: #1A73E8;
}
QStatusBar {
    background-color: #F8F9FA;
    border-top: 1px solid #DADCE0;
    color: #5F6368;
    font-size: 12px;
}
QScrollBar:vertical {
    background: #F8F9FA;
    width: 12px;
    border: none;
}
QScrollBar::handle:vertical {
    background: #BDC1C6;
    border-radius: 6px;
    min-height: 30px;
}
QScrollBar::handle:vertical:hover {
    background: #9AA0A6;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    height: 0;
}
QScrollBar:horizontal {
    background: #F8F9FA;
    height: 12px;
    border: none;
}
QScrollBar::handle:horizontal {
    background: #BDC1C6;
    border-radius: 6px;
    min-width: 30px;
}
QScrollBar::handle:horizontal:hover {
    background: #9AA0A6;
}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
    width: 0;
}
#formula_label {
    background-color: #F1F3F4;
    border: 1px solid #DADCE0;
    border-radius: 3px;
    padding: 2px 6px;
    font-family: Arial;
    font-size: 12px;
    font-weight: bold;
    color: #5F6368;
}
#formula_edit {
    background-color: #FFFFFF;
    border: 1px solid #DADCE0;
    border-radius: 3px;
    padding: 2px 6px;
    font-family: Arial;
    font-size: 13px;
    color: #202124;
}
#formula_edit:focus {
    border: 2px solid #1A73E8;
}
"""


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    app = QApplication(sys.argv)
    app.setApplicationName("SheetEdit")
    app.setStyle("Fusion")
    app.setStyleSheet(LIGHT_STYLESHEET)

    filepath = sys.argv[1] if len(sys.argv) > 1 else None
    win = SheetEditWindow(filepath)
    win.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()
